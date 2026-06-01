"""
physical_plausibility.py
============================
RQ2 – Physical Plausibility Analysis for BayScen paper.

Measures the *physical violation rate*: the proportion of generated
scenarios containing at least one parameter combination that is
physically impossible in the real world.

Constraints are taken from Hao et al. (2023) (BridgeGen) and grounded
in ISO 34503:2023 (§9.3.7, §10.3).

CONSTRAINTS
-----------
All feature_* columns are in CARLA's native scale:
  Cloudiness, Precipitation, PrecipitationDeposits, WindIntensity,
  FogDensity, FogDistance, Wetness  → [0, 100]  (integer)
  RoadFriction                      → [0.1, 1.0] (float)

  C1   P > 20 ⟹  D > 0               (meaningful precipitation causes deposits)
  C2   P > 20 ⟹  W > 0               (meaningful precipitation causes wetness)
  C3a  W < 40  ⟹  F ≤ 1 − W/200      (wetness reduces friction, low regime)
  C3b  W ≥ 40  ⟹  F ≤ 0.6            (wetness reduces friction, high regime)
  C4   |L − (100 − G)| ≤ ε=10        (fog density determines fog distance)
  C5   N ≥ 60  ⟹  G ≤ 40             (high wind disperses fog)
  C6   P > 20  ⟹  C > 0              (rain requires cloud cover)

Changes from previous version
------------------------------
  - C1: threshold raised from P > 0  to P > 20
  - C2: threshold raised from P > 0  to P > 20
  - C3: split into two regimes (C3a low, C3b high) per BridgeGen Eq.(1)
        C3a: W < 40  ⟹  F ≤ 1 − W/200   [NEW]
        C3b: W ≥ 40  ⟹  F ≤ 0.6          [unchanged logic, renamed]
  - C5: threshold changed from N > 60, G < 40  to  N ≥ 60, G ≤ 40
  - C6: threshold changed from P > 0, C > 20   to  P > 20, C > 0

Variable abbreviations:
  P = feature_Precipitation       D = feature_PrecipitationDeposits
  W = feature_Wetness             F = feature_RoadFriction
  G = feature_FogDensity          L = feature_FogDistance
  N = feature_WindIntensity       C = feature_Cloudiness

FILE STRUCTURE EXPECTED
-----------------------
simulation results/
  Interfuser/
    Scenario 1/
      avfuzzer_scenario1_interfuser_run1.csv
      ...
    Scenario 2/ ...
    Scenario 3/ ...
  Modular/ ...

OUTPUT  →  ./results/rq2/physical_plausibility.xlsx
  Sheet per SUT  +  Summary pivot sheets

USAGE
-----
  python scripts/physical_plausibility.py --results "simulation results"
  python scripts/physical_plausibility.py --results "simulation results" --output results/rq2/physical_plausibility.xlsx
"""

import argparse
import glob
import logging
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

SUTS = ["Interfuser", "Modular"]

SCENARIO_FOLDERS = {
    1: "Scenario 1",
    2: "Scenario 2",
    3: "Scenario 3",
}

BASELINE_ORDER = [
    "random", "sitcov", "pict_2way", "pict_3way",
    "ctbc", "avfuzzer", "bayscen_common", "bayscen",
]

BASELINE_LABELS = {
    "random":         "Random",
    "sitcov":         "SitCov",
    "pict_2way":      "PICT 2-way",
    "pict_3way":      "PICT 3-way",
    "ctbc":           "CTBC",
    "avfuzzer":       "AvFuzzer",
    "bayscen_common": "BayScen-Common",
    "bayscen":        "BayScen",
}

PREFIX_MAP = {
    "bayscen_common": "bayscen_common",
    "bayscen":        "bayscen",
    "avfuzzer":       "avfuzzer",
    "random":         "random",
    "sitcov":         "sitcov",
    "pict_2way":      "pict_2way",
    "pict_2w":        "pict_2way",
    "pict_3way":      "pict_3way",
    "pict_3w":        "pict_3way",
    "ctbc":           "ctbc",
}

# Constraint parameter ε for C4
C4_EPSILON = 10.0

# Precipitation threshold for C1, C2, C6 — meaningful rainfall only
P_THRESHOLD = 20.0

# Constraint definitions: (id, description, formal_rule)
# C3 is split into two sub-constraints (C3a, C3b) per BridgeGen Eq.(1)
CONSTRAINTS = [
    ("C1",  "Precipitation causes deposits",        "P > 20 ⟹ D > 0"),
    ("C2",  "Precipitation causes wetness",         "P > 20 ⟹ W > 0"),
    ("C3a", "Wetness reduces friction (low regime)", "W < 40 ⟹ F ≤ 1 − W/200"),
    ("C3b", "Wetness reduces friction (high regime)","W ≥ 40 ⟹ F ≤ 0.6"),
    ("C4",  "Fog density determines distance",      "|L − (100−G)| ≤ 10"),
    ("C5",  "High wind disperses fog",              "N ≥ 60 ⟹ G ≤ 40"),
    ("C6",  "Rain requires cloud cover",            "P > 20 ⟹ C > 0"),
]

# Ordered constraint IDs — used wherever we iterate over constraints
CONSTRAINT_IDS = [c[0] for c in CONSTRAINTS]

# ──────────────────────────────────────────────────────────────────────────────
# Filename parsing  (shared with criticality.py)
# ──────────────────────────────────────────────────────────────────────────────

def parse_filename(filename: str):
    stem = Path(filename).stem
    run_match = re.search(r"_run(\d+)$", stem, re.IGNORECASE)
    if not run_match:
        return None, None
    run_num = int(run_match.group(1))
    prefix_part = stem[: run_match.start()].lower()
    prefix_part = re.sub(r"_scenario\d+_\w+$", "", prefix_part)
    for key in sorted(PREFIX_MAP.keys(), key=len, reverse=True):
        if prefix_part == key or prefix_part.startswith(key):
            return PREFIX_MAP[key], run_num
    return prefix_part, run_num


# ──────────────────────────────────────────────────────────────────────────────
# Physical plausibility constraints
# ──────────────────────────────────────────────────────────────────────────────

def _col(df: pd.DataFrame, *candidates: str) -> pd.Series:
    """
    Return the first column from *candidates* that exists in df (case-insensitive).
    Raises KeyError with a clear message if none found.
    """
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return df[lower_map[cand.lower()]]
    raise KeyError(
        f"None of {candidates} found in columns: {df.columns.tolist()}"
    )


def check_constraints(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluate all physical plausibility constraints on every row.

    Returns a boolean DataFrame with one column per constraint (True = violated).
    Columns: C1, C2, C3a, C3b, C4, C5, C6.

    Constraint definitions
    ----------------------
    C1  : P > 20 AND D == 0           precipitation without deposits
    C2  : P > 20 AND W == 0           precipitation without wetness
    C3a : W < 40  AND F > 1 − W/200   friction too low for dry-ish road
                                       (low-wetness regime, BridgeGen Eq.1)
    C3b : W >= 40 AND F > 0.6         friction too high for wet road
                                       (high-wetness regime, BridgeGen Eq.1)
    C4  : |L − (100 − G)| > ε         fog distance inconsistent with density
    C5  : N >= 60 AND G > 40          high wind with dense fog
    C6  : P > 20  AND C == 0          rain without any cloud cover
    """
    P = _col(df, "feature_Precipitation")
    D = _col(df, "feature_PrecipitationDeposits")
    W = _col(df, "feature_Wetness")
    F = _col(df, "feature_RoadFriction")
    G = _col(df, "feature_FogDensity")
    L = _col(df, "feature_FogDistance")
    N = _col(df, "feature_WindIntensity")
    C = _col(df, "feature_Cloudiness")

    violations = pd.DataFrame(index=df.index)

    # C1: meaningful precipitation (P > 20) without deposits
    violations["C1"] = (P > P_THRESHOLD) & (D == 0)

    # C2: meaningful precipitation (P > 20) without wetness
    violations["C2"] = (P > P_THRESHOLD) & (W == 0)

    # C3a: low-wetness regime — friction must follow F ≤ 1 − W/200
    #      Only evaluated when W < 40; violated if F exceeds the formula value.
    #      Note: 1 − W/200 is on [0,1] scale; F is already in [0.1, 1.0].
    friction_limit_low = 1.0 - W / 200.0
    violations["C3a"] = (W < 40) & (F > friction_limit_low)

    # C3b: high-wetness regime — friction must not exceed 0.6
    violations["C3b"] = (W >= 40) & (F > 0.6)

    # C4: fog distance must satisfy |L − (100 − G)| ≤ ε
    violations["C4"] = (L - (100 - G)).abs() > C4_EPSILON

    # C5: high wind (N ≥ 60) must not co-occur with dense fog (G > 40)
    violations["C5"] = (N >= 60) & (G > 40)

    # C6: meaningful precipitation (P > 20) requires some cloud cover (C > 0)
    violations["C6"] = (P > P_THRESHOLD) & (C == 0)

    return violations


def compute_plausibility_metrics(df: pd.DataFrame) -> dict:
    """
    Compute per-constraint violation rates and the overall physical violation
    rate (at least one constraint violated per scenario row).

    Returns a dict with:
      n_scenarios
      any_violation_count      – rows with ≥ 1 constraint violated
      physical_violation_rate  – any_violation_count / n_scenarios
      physical_violation_pct   – rate × 100
      C1_count … C6_count      – per-constraint violation counts
      C1_rate  … C6_rate       – per-constraint violation rates
      C1_pct   … C6_pct        – per-constraint violation percentages
      (same for C3a, C3b)
    """
    viol = check_constraints(df)
    n    = len(df)

    any_viol   = viol.any(axis=1)
    any_count  = int(any_viol.sum())
    any_rate   = any_count / n if n > 0 else 0.0

    metrics = {
        "n_scenarios":             n,
        "any_violation_count":     any_count,
        "physical_violation_rate": any_rate,
        "physical_violation_pct":  any_rate * 100,
    }

    for c in CONSTRAINT_IDS:
        cnt  = int(viol[c].sum())
        rate = cnt / n if n > 0 else 0.0
        metrics[f"{c}_count"] = cnt
        metrics[f"{c}_rate"]  = rate
        metrics[f"{c}_pct"]   = rate * 100

    return metrics


# ──────────────────────────────────────────────────────────────────────────────
# File indexing + loading
# ──────────────────────────────────────────────────────────────────────────────

def index_files(results_dir: Path) -> dict:
    file_index: dict = {}
    for sut in SUTS:
        for scen_num, scen_folder in SCENARIO_FOLDERS.items():
            folder = results_dir / sut / scen_folder
            if not folder.is_dir():
                log.warning("Folder not found: %s", folder)
                continue
            for fpath in glob.glob(str(folder / "*.csv")):
                baseline, run_num = parse_filename(fpath)
                if baseline is None:
                    log.warning("Cannot parse filename: %s", fpath)
                    continue
                key = (sut, scen_num, baseline)
                file_index.setdefault(key, {})[run_num] = Path(fpath)
    return file_index


def load_csv(fpath: Path) -> pd.DataFrame:
    """Load a per-run CSV and normalise column names."""
    df = pd.read_csv(fpath)
    df.columns = df.columns.str.strip()
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Main analysis loop
# ──────────────────────────────────────────────────────────────────────────────

def compute_all(file_index: dict) -> pd.DataFrame:
    """
    For each (sut, scenario, baseline) group, compute per-run and
    aggregated physical plausibility metrics.
    """
    records = []

    all_groups = [
        ((sut, scen, bl), runs)
        for (sut, scen, bl), runs in sorted(file_index.items())
        if len(runs) >= 3
    ]

    for (sut, scen, baseline), run_files in all_groups:
        label = BASELINE_LABELS.get(baseline, baseline)
        log.info("%-12s | Scen %d | %s", sut, scen, label)

        row = {"sut": sut, "scenario": scen, "baseline": baseline}

        per_run_metrics = []
        for r in (1, 2, 3):
            fpath = run_files.get(r)
            if fpath is None:
                log.warning("  Run %d missing", r)
                continue
            try:
                df      = load_csv(fpath)
                metrics = compute_plausibility_metrics(df)
                per_run_metrics.append(metrics)

                # Store per-run values
                row[f"violation_rate_run{r}"] = metrics["physical_violation_rate"]
                row[f"violation_pct_run{r}"]  = metrics["physical_violation_pct"]
                for c in CONSTRAINT_IDS:
                    row[f"{c}_pct_run{r}"] = metrics[f"{c}_pct"]

                log.info(
                    "  run%d  n=%d  viol=%.1f%%  "
                    "(C1=%.0f%% C2=%.0f%% C3a=%.0f%% C3b=%.0f%% "
                    "C4=%.0f%% C5=%.0f%% C6=%.0f%%)",
                    r, metrics["n_scenarios"], metrics["physical_violation_pct"],
                    metrics["C1_pct"], metrics["C2_pct"],
                    metrics["C3a_pct"], metrics["C3b_pct"],
                    metrics["C4_pct"], metrics["C5_pct"], metrics["C6_pct"],
                )
            except Exception as exc:
                log.error("  run%d FAILED: %s", r, exc)

        if not per_run_metrics:
            continue

        # Aggregate across runs (mean ± std)
        for key in ["physical_violation_rate", "physical_violation_pct",
                    "n_scenarios"]:
            vals = [m[key] for m in per_run_metrics]
            row[f"{key}_mean"] = float(np.mean(vals))
            row[f"{key}_std"]  = float(np.std(vals))

        for c in CONSTRAINT_IDS:
            pct_vals = [m[f"{c}_pct"] for m in per_run_metrics]
            row[f"{c}_pct_mean"] = float(np.mean(pct_vals))
            row[f"{c}_pct_std"]  = float(np.std(pct_vals))

        row["n_scenarios"] = int(np.mean([m["n_scenarios"] for m in per_run_metrics]))

        log.info(
            "  ✔  mean violation rate = %.2f%% ± %.2f%%",
            row["physical_violation_pct_mean"],
            row["physical_violation_pct_std"],
        )
        records.append(row)

    if not records:
        raise RuntimeError("No data loaded. Check folder structure and filenames.")

    results = pd.DataFrame(records)

    # Sort in canonical baseline order
    order_map = {b: i for i, b in enumerate(BASELINE_ORDER)}
    results["_ord"] = results["baseline"].map(lambda b: order_map.get(b, 999))
    results = (results
               .sort_values(["sut", "scenario", "_ord"])
               .drop(columns="_ord")
               .reset_index(drop=True))
    return results


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def _style_sheet(ws, bayscen_row_indices: list):
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    bscn_fill = PatternFill("solid", fgColor="C8E6C9")
    alt_fill  = PatternFill("solid", fgColor="EEF2F7")
    bw   = Font(bold=True, color="FFFFFF")
    bb   = Font(bold=True)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDBDBD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, bw, ctr, brd
    ws.row_dimensions[1].height = 36

    for ri in range(2, ws.max_row + 1):
        is_bs = (ri - 1) in bayscen_row_indices
        for cell in ws[ri]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            if is_bs:
                cell.fill, cell.font = bscn_fill, bb
            elif ri % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 24)
    ws.freeze_panes = "C2"


def export_excel(results: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Sheet 1: per-SUT detailed view ────────────────────────────────────────
    # Columns: scenario | baseline | n_scenarios
    #          violation_pct_run1/2/3 | violation_pct_mean | violation_pct_std
    #          C1_pct_mean … C6_pct_mean  (including C3a, C3b)

    detail_cols = (
        ["scenario", "baseline", "n_scenarios"]
        + [f"violation_pct_run{r}" for r in (1, 2, 3)]
        + ["physical_violation_pct_mean", "physical_violation_pct_std"]
        + [f"{c}_pct_mean" for c in CONSTRAINT_IDS]
        + [f"{c}_pct_std"  for c in CONSTRAINT_IDS]
    )

    header_map = {
        "scenario":                      "Scenario",
        "baseline":                      "Baseline",
        "n_scenarios":                   "N Scenarios",
        "violation_pct_run1":            "Viol. % run1",
        "violation_pct_run2":            "Viol. % run2",
        "violation_pct_run3":            "Viol. % run3",
        "physical_violation_pct_mean":   "Viol. % mean",
        "physical_violation_pct_std":    "Viol. % std",
        **{f"{c}_pct_mean": f"{c} % mean" for c in CONSTRAINT_IDS},
        **{f"{c}_pct_std":  f"{c} % std"  for c in CONSTRAINT_IDS},
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for sut in SUTS:
            sub = results[results["sut"] == sut].copy()
            sub["baseline"] = sub["baseline"].map(lambda b: BASELINE_LABELS.get(b, b))
            existing = [c for c in detail_cols if c in sub.columns]
            out = sub[existing].copy()
            out.columns = [header_map.get(c, c) for c in existing]
            out = out.round(2)
            out.to_excel(writer, sheet_name=sut, index=False)

        # ── Summary pivot: violation_pct_mean ─────────────────────────────────
        pivot_viol = results.pivot_table(
            index="baseline", columns=["sut", "scenario"],
            values="physical_violation_pct_mean", aggfunc="first",
        )
        known = [b for b in BASELINE_ORDER if b in pivot_viol.index]
        pivot_viol = pivot_viol.reindex(
            known + [b for b in pivot_viol.index if b not in known]
        )
        pivot_viol.index = [BASELINE_LABELS.get(b, b) for b in pivot_viol.index]
        pivot_viol.columns = [f"{s}/S{sc}" for s, sc in pivot_viol.columns]
        pivot_viol.round(2).to_excel(writer, sheet_name="Violation Rate (%) Summary")

        # ── Per-constraint summary (mean across runs, averaged over scenarios) ─
        for c_id, c_desc, c_rule in CONSTRAINTS:
            col = f"{c_id}_pct_mean"
            if col not in results.columns:
                continue
            pivot_c = results.pivot_table(
                index="baseline", columns=["sut", "scenario"],
                values=col, aggfunc="first",
            )
            known_c = [b for b in BASELINE_ORDER if b in pivot_c.index]
            pivot_c = pivot_c.reindex(
                known_c + [b for b in pivot_c.index if b not in known_c]
            )
            pivot_c.index = [BASELINE_LABELS.get(b, b) for b in pivot_c.index]
            pivot_c.columns = [f"{s}/S{sc}" for s, sc in pivot_c.columns]
            sheet_name = f"{c_id} ({c_desc[:20]})"[:31]
            pivot_c.round(2).to_excel(writer, sheet_name=sheet_name)

        # ── Constraint legend sheet ────────────────────────────────────────────
        legend_rows = []
        grounding_map = {
            "C1":  "ISO 34503 §9.3.7",
            "C2":  "ISO 34503 §9.3.7",
            "C3a": "BridgeGen Eq. (1)",
            "C3b": "BridgeGen Eq. (1)",
            "C4":  "BridgeGen Eq. (2)",
            "C5":  "BridgeGen §III-B",
            "C6":  "BridgeGen §III-B",
        }
        for cid, desc, rule in CONSTRAINTS:
            legend_rows.append({
                "Constraint":            cid,
                "Physical Relationship": desc,
                "Formal Rule":           rule,
                "Grounding":             grounding_map.get(cid, ""),
                "Change from v1":        _changelog(cid),
            })
        legend = pd.DataFrame(legend_rows)
        legend.to_excel(writer, sheet_name="Constraints", index=False)

    # ── Post-process styling ───────────────────────────────────────────────────
    wb = load_workbook(output_path)

    for sut in SUTS:
        if sut not in wb.sheetnames:
            continue
        ws = wb[sut]
        sub = results[results["sut"] == sut].reset_index(drop=True)
        bs_rows = [i + 1 for i, b in enumerate(sub["baseline"]) if b == "bayscen"]
        _style_sheet(ws, bs_rows)

    # Light header styling on summary/pivot sheets
    for sh in wb.sheetnames:
        if sh in SUTS:
            continue
        ws = wb[sh]
        hf = PatternFill("solid", fgColor="1F4E79")
        bw = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill, cell.font = hf, bw
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max((len(str(c.value or "")) for c in col), default=8) + 4
            )

    wb.save(output_path)
    log.info("Saved → %s", output_path)


def _changelog(cid: str) -> str:
    """Human-readable summary of what changed per constraint vs. old version."""
    changes = {
        "C1":  "Threshold raised: P > 0 → P > 20",
        "C2":  "Threshold raised: P > 0 → P > 20",
        "C3a": "NEW — low-wetness friction regime: W < 40 ⟹ F ≤ 1 − W/200",
        "C3b": "Renamed from C3; logic unchanged: W ≥ 40 ⟹ F ≤ 0.6",
        "C4":  "Unchanged",
        "C5":  "Strict inequalities → non-strict: N > 60, G < 40 → N ≥ 60, G ≤ 40",
        "C6":  "Thresholds changed: P > 0, C > 20 → P > 20, C > 0",
    }
    return changes.get(cid, "")


# ──────────────────────────────────────────────────────────────────────────────
# Console summary
# ──────────────────────────────────────────────────────────────────────────────

def print_summary(results: pd.DataFrame) -> None:
    print(f"\n{'='*70}")
    print("  RQ2 – Physical Violation Rate (%) [mean across 3 runs]")
    print(f"{'='*70}")
    pivot = results.pivot_table(
        index="baseline", columns=["sut", "scenario"],
        values="physical_violation_pct_mean", aggfunc="first",
    )
    known = [b for b in BASELINE_ORDER if b in pivot.index]
    pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
    pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
    pivot.columns = [f"{s}/S{sc}" for s, sc in pivot.columns]
    print(pivot.round(2).to_string())

    print(f"\n{'='*70}")
    print("  Per-Constraint Violation Rate (%) [mean across runs & scenarios]")
    print(f"{'='*70}")
    rows = []
    for c_id, c_desc, c_rule in CONSTRAINTS:
        col = f"{c_id}_pct_mean"
        if col not in results.columns:
            continue
        avg_per_baseline = (
            results.groupby("baseline")[col]
            .mean()
            .reindex([b for b in BASELINE_ORDER if b in results["baseline"].values])
        )
        avg_per_baseline.index = [BASELINE_LABELS.get(b, b) for b in avg_per_baseline.index]
        rows.append(avg_per_baseline.rename(f"{c_id}: {c_desc[:30]}"))

    if rows:
        df_c = pd.concat(rows, axis=1).T
        print(df_c.round(2).to_string())

    print(f"\n{'='*70}")
    print("  Constraint definitions")
    print(f"{'='*70}")
    for c_id, c_desc, c_rule in CONSTRAINTS:
        print(f"  {c_id:<4}  {c_rule:<30}  {c_desc}")


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="RQ2 – Physical Plausibility Analysis for BayScen (updated constraints).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--results", type=Path, default=Path("simulation results"),
        help="Root folder containing Interfuser/ and Modular/ sub-folders.",
    )
    p.add_argument(
        "--output", type=Path,
        default=Path("results/rq2/physical_plausibility.xlsx"),
        help="Output Excel file path.",
    )
    p.add_argument(
        "--epsilon", type=float, default=10.0,
        help="Tolerance ε for constraint C4: |L − (100−G)| ≤ ε.",
    )
    p.add_argument(
        "--precip-threshold", type=float, default=20.0,
        help="Precipitation threshold for C1, C2, C6 (default 20, i.e. P > 20).",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    # Apply CLI overrides
    global C4_EPSILON, P_THRESHOLD
    C4_EPSILON  = args.epsilon
    P_THRESHOLD = args.precip_threshold

    log.info("=" * 65)
    log.info("RQ2 – Physical Plausibility Analysis (updated constraints)")
    log.info("=" * 65)
    log.info("  Results dir        : %s", args.results)
    log.info("  Output             : %s", args.output)
    log.info("  C4 epsilon         : %.1f", C4_EPSILON)
    log.info("  Precip threshold   : %.1f (C1, C2, C6)", P_THRESHOLD)
    log.info("  Constraints        : %s", ", ".join(CONSTRAINT_IDS))
    log.info("=" * 65)

    if not args.results.is_dir():
        log.error("Results directory not found: %s", args.results)
        raise SystemExit(1)

    log.info("\nIndexing CSV files …")
    file_index = index_files(args.results)
    n_groups = sum(1 for runs in file_index.values() if len(runs) >= 3)
    log.info("Found %d complete groups (≥ 3 runs).\n", n_groups)

    if n_groups == 0:
        log.error("No complete groups found. Check folder structure and filenames.")
        raise SystemExit(1)

    log.info("Computing physical plausibility constraints …\n")
    results = compute_all(file_index)

    log.info("\nExporting Excel …")
    export_excel(results, args.output)

    print_summary(results)

    print(f"\nAll RQ2 outputs saved to: {args.output}\n")


if __name__ == "__main__":
    main()
