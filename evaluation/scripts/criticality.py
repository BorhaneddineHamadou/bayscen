"""
criticality.py  –  RQ1 Criticality Analysis for BayScen paper
==============================================================

ACTUAL FILE STRUCTURE
---------------------
simulation results/
  Interfuser/
    Scenario 1/
      avfuzzer_scenario1_interfuser_run1.csv
      avfuzzer_scenario1_interfuser_run2.csv
      avfuzzer_scenario1_interfuser_run3.csv
      bayscen_scenario1_interfuser_run1.csv
      ...
    Scenario 2/ ...
    Scenario 3/ ...
  Modular/
    Scenario 1/ ...
    ...

Each per-run CSV contains columns:
  Collision   – bool (True / False)
  MinTTC      – float; 9999.0 = no conflict; 0.0 = collision
  algo_safety – int  (1 if MinTTC < 0.5 OR Collision == True)
  + feature_* columns (ignored here)

A scenario row is safety-critical if its MEAN MinTTC across the 3 runs
< 0.5 s (sentinel 9999 excluded from the mean).

OUTPUTS  ->  ./results/rq1/
  rq1_raw_data.xlsx          per-baseline metrics, one sheet per SUT
  rq1_summary.xlsx           pivot table ready for the paper
  rq1_collision_rate.png     grouped bar chart per SUT
  rq1_ttc_critical_rate.png  grouped bar chart per SUT
  rq1_combined.png           2x2 paper-ready figure
"""

import os
import re
import glob
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ------------------------------------------------------------------------------
# CONFIGURATION  (edit paths here if needed)
# ------------------------------------------------------------------------------

BASE_DIR      = "./simulation results"
RESULTS_DIR   = "./results/rq1"
TTC_THRESHOLD = 0.5        # seconds – safety-critical threshold (paper §III-E)
TTC_EPSILON   = 0.001      # absorbs CSV rounding: 0.4999 stored as 0.500 with 3dp
TTC_SENTINEL  = 9999.0     # means "no conflict" – excluded from TTC statistics

SUTS = ["Interfuser", "Modular"]

SCENARIO_FOLDERS = {
    1: "Scenario 1",
    2: "Scenario 2",
    3: "Scenario 3",
}

# Canonical display order for baselines
BASELINE_ORDER = [
    "random",
    "sitcov",
    "pict_2way",
    "pict_3way",
    "ctbc",
    "avfuzzer",
    "bayscen_common",
    "bayscen",
]

BASELINE_LABELS = {
    "random":         "Random",
    "sitcov":         "SitCov",
    "pict_2way":      "PICT 2-way",
    "pict_3way":      "PICT 3-way",
    "ctbc":           "CTBC",
    "avfuzzer":       "AvFuzzer",
    "bayscen_common": "BayScen-Common",
    "bayscen":        "BayScen",
}

# Filename prefix -> canonical key (longest match wins, so bayscen_common before bayscen)
PREFIX_MAP = {
    "random":         "random",
    "sitcov":         "sitcov",
    "pict_2w":        "pict_2way",
    "pict_2way":      "pict_2way",
    "pict_3w":        "pict_3way",
    "pict_3way":      "pict_3way",
    "ctbc":           "ctbc",
    "avfuzzer":       "avfuzzer",
    "bayscen_common": "bayscen_common",
    "bayscen":        "bayscen",
}

BASELINE_COLORS = {
    "random":         "#9E9E9E",
    "sitcov":         "#78909C",
    "pict_2way":      "#4FC3F7",
    "pict_3way":      "#0288D1",
    "ctbc":           "#F4A261",
    "avfuzzer":       "#E63946",
    "bayscen_common": "#8BC34A",
    "bayscen":        "#2E7D32",
}

# ------------------------------------------------------------------------------
# FILENAME PARSING
# ------------------------------------------------------------------------------

def parse_filename(filename):
    """
    Extract (baseline_key, run_number) from filenames like:
      avfuzzer_scenario2_modular_run1.csv
      PICT_2w_scenario3_modular_run2.csv
      bayscen_common_scenario1_interfuser_run3.csv

    Returns (baseline_key, run_int) or (None, None) if not parseable.
    """
    stem = os.path.splitext(os.path.basename(filename))[0]

    # Extract run number from the end
    run_match = re.search(r"_run(\d+)$", stem, re.IGNORECASE)
    if not run_match:
        return None, None
    run_num = int(run_match.group(1))

    # Remove _runN suffix, then _scenarioN_<sut>
    prefix_part = stem[: run_match.start()].lower()
    prefix_part = re.sub(r"_scenario\d+_\w+$", "", prefix_part)

    # Match against PREFIX_MAP – longest key first (bayscen_common before bayscen)
    for key in sorted(PREFIX_MAP.keys(), key=len, reverse=True):
        if prefix_part == key or prefix_part.startswith(key + "_") or prefix_part == key:
            return PREFIX_MAP[key], run_num

    # Unknown baseline – keep raw name so it still appears in output
    return prefix_part, run_num


# ------------------------------------------------------------------------------
# DATA LOADING & METRIC COMPUTATION
# ------------------------------------------------------------------------------

def load_run_csv(filepath):
    """
    Load one per-run CSV and normalise column names.

    Handles all observed column name variants:
      Collision / collision            -> 'collision'
      MinTTC    / min_ttc / minttc    -> 'minttc'
      algo_safety                     -> 'algo_safety'
    """
    df = pd.read_csv(filepath)

    # Strip whitespace, then build a normalisation map
    df.columns = [c.strip() for c in df.columns]

    rename = {}
    for c in df.columns:
        cl = c.lower()
        # Collision variants
        if cl == "collision":
            rename[c] = "collision"
        # TTC variants: MinTTC, min_ttc, minttc, min_TTC, …
        elif cl.replace("_", "") == "minttc":
            rename[c] = "minttc"
        else:
            rename[c] = cl          # lowercase everything else

    df = df.rename(columns=rename)

    if "collision" not in df.columns:
        raise ValueError(f"Missing 'Collision' column in {filepath}")
    if "minttc" not in df.columns:
        raise ValueError(f"Missing 'MinTTC' / 'min_ttc' column in {filepath}")

    df["collision"] = df["collision"].astype(bool).astype(int)
    df["minttc"]    = pd.to_numeric(df["minttc"], errors="coerce")
    return df


def compute_metrics_from_runs(run_dfs):
    """
    Aggregate RQ1 metrics across 3 aligned run DataFrames.

    collision_count     – rows where ANY run has Collision == True
    ttc_critical_count  – rows where MEAN MinTTC (excl. 9999) <= TTC_THRESHOLD + TTC_EPSILON
    """
    n = len(run_dfs[0])

    # Collision: True in any run
    col_mat = np.stack([df["collision"].values for df in run_dfs], axis=1)
    coll_any = col_mat.max(axis=1)
    collision_count = int(coll_any.sum())
    collision_rate  = collision_count / n if n > 0 else 0.0

    # TTC: exclude sentinels, mean across runs
    ttc_mat = np.stack([df["minttc"].values for df in run_dfs], axis=1).astype(float)
    ttc_mat[ttc_mat >= TTC_SENTINEL] = np.nan
    mean_ttc = np.nanmean(ttc_mat, axis=1)
    # Use <= (TTC_THRESHOLD + TTC_EPSILON) to catch values like 0.500 that
    # were originally <0.5 in the simulator but rounded up by CSV float_format.
    ttc_crit_count = int(np.nansum(mean_ttc <= TTC_THRESHOLD + TTC_EPSILON))
    ttc_crit_rate  = ttc_crit_count / n if n > 0 else 0.0
    mean_min_ttc   = float(np.nanmean(mean_ttc))

    # algo_safety union (if present)
    algo_safety_count = None
    if "algo_safety" in run_dfs[0].columns:
        as_mat = np.stack([df["algo_safety"].values for df in run_dfs], axis=1)
        algo_safety_count = int(as_mat.max(axis=1).sum())

    return {
        "n_scenarios":              n,
        "collision_count":          collision_count,
        "collision_rate":           collision_rate,
        "collision_rate_per100":    collision_rate * 100,
        "ttc_critical_count":       ttc_crit_count,
        "ttc_critical_rate":        ttc_crit_rate,
        "ttc_critical_rate_per100": ttc_crit_rate * 100,
        "mean_min_ttc":             mean_min_ttc,
        "algo_safety_count":        algo_safety_count,
    }


# ------------------------------------------------------------------------------
# DATA COLLECTION
# ------------------------------------------------------------------------------

def collect_all_results():
    """
    Walk BASE_DIR, group CSVs by (sut, scenario, baseline),
    load the 3 runs, compute metrics, return a long-format DataFrame.
    """
    file_index = {}   # (sut, scen_num, baseline) -> {run_num: filepath}

    for sut in SUTS:
        for scen_num, scen_folder in SCENARIO_FOLDERS.items():
            folder = os.path.join(BASE_DIR, sut, scen_folder)
            if not os.path.isdir(folder):
                print(f"  [WARN] Folder not found: {folder}")
                continue
            for fpath in glob.glob(os.path.join(folder, "*.csv")):
                baseline, run_num = parse_filename(fpath)
                if baseline is None:
                    print(f"  [SKIP] Cannot parse filename: {fpath}")
                    continue
                key = (sut, scen_num, baseline)
                file_index.setdefault(key, {})[run_num] = fpath

    if not file_index:
        raise RuntimeError(
            "No files found. Check BASE_DIR and that filenames contain '_runN'."
        )

    records = []
    for (sut, scen_num, baseline), run_files in sorted(file_index.items()):
        runs_found = sorted(run_files.keys())

        if len(runs_found) < 3:
            files_listed = [os.path.basename(run_files[r]) for r in runs_found]
            print(f"  [WARN] {sut} | Scen {scen_num} | {baseline}: "
                  f"only {len(runs_found)} run(s) found {runs_found} – skipping. "
                  f"Files: {files_listed}  "
                  f"(check for filename typos, e.g. 'ranom' instead of 'random')")
            continue

        try:
            run_dfs = [load_run_csv(run_files[r]) for r in [1, 2, 3]]
            metrics = compute_metrics_from_runs(run_dfs)
            records.append({"sut": sut, "scenario": scen_num,
                             "baseline": baseline, **metrics})
            print(f"  OK  {sut:12s} | Scen {scen_num} | {baseline:20s} "
                  f"| n={metrics['n_scenarios']:4d} "
                  f"| coll={metrics['collision_rate']:.3f} "
                  f"| ttc={metrics['ttc_critical_rate']:.3f}")
        except Exception as exc:
            print(f"  [ERR] {sut} | Scen {scen_num} | {baseline}: {exc}")

    if not records:
        raise RuntimeError("No valid data loaded. Check file structure and column names.")

    results = pd.DataFrame(records)
    order_map = {b: i for i, b in enumerate(BASELINE_ORDER)}
    results["_ord"] = results["baseline"].map(lambda b: order_map.get(b, 999))
    results = (results.sort_values(["sut", "scenario", "_ord"])
                      .drop(columns="_ord").reset_index(drop=True))
    return results


# ------------------------------------------------------------------------------
# EXCEL EXPORT
# ------------------------------------------------------------------------------

def _style_sheet(ws, col_labels, data_rows):
    hdr  = PatternFill("solid", fgColor="1F4E79")
    bscn = PatternFill("solid", fgColor="C8E6C9")
    alt  = PatternFill("solid", fgColor="EEF2F7")
    bw   = Font(bold=True, color="FFFFFF")
    bb   = Font(bold=True)
    ctr  = Alignment(horizontal="center", vertical="center")
    th   = Side(style="thin", color="BDBDBD")
    brd  = Border(left=th, right=th, top=th, bottom=th)

    for ci, lbl in enumerate(col_labels, 1):
        c = ws.cell(row=1, column=ci, value=lbl)
        c.fill, c.font, c.alignment, c.border = hdr, bw, ctr, brd

    for ri, row in enumerate(data_rows, 2):
        is_bs = str(row.get("baseline", "")).lower() == "bayscen"
        for ci, key in enumerate(col_labels, 1):
            raw = row.get(key, "")
            val = round(raw, 4) if isinstance(raw, float) else raw
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment, c.border = ctr, brd
            if is_bs:
                c.fill, c.font = bscn, bb
            elif ri % 2 == 0:
                c.fill = alt

    for ci, lbl in enumerate(col_labels, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(lbl) + 4, 12)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def export_raw_excel(results, out_path):
    keys = ["scenario", "baseline", "n_scenarios",
            "collision_count", "collision_rate", "collision_rate_per100",
            "ttc_critical_count", "ttc_critical_rate", "ttc_critical_rate_per100",
            "mean_min_ttc", "algo_safety_count"]
    labels = ["Scenario", "Baseline", "N Scenarios",
              "Collision Count", "Collision Rate", "Collision Rate /100",
              "TTC-Critical Count", "TTC-Critical Rate", "TTC-Critical Rate /100",
              "Mean Min TTC (s)", "Algo-Safety Count"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        for sut in SUTS:
            sub = results[results["sut"] == sut].copy()
            sub["baseline"] = sub["baseline"].map(lambda b: BASELINE_LABELS.get(b, b))
            sub[keys].to_excel(w, sheet_name=sut, index=False)

    wb = load_workbook(out_path)
    for sut in SUTS:
        if sut not in wb.sheetnames:
            continue
        ws = wb[sut]
        sub = results[results["sut"] == sut].copy()
        sub["baseline"] = sub["baseline"].map(lambda b: BASELINE_LABELS.get(b, b))
        # Build rows with label-keyed dicts so _style_sheet can find values
        key_to_label = dict(zip(keys, labels))
        rows_labeled = [
            {key_to_label[k]: row[k] for k in keys}
            for row in sub[keys].to_dict("records")
        ]
        _style_sheet(ws, labels, rows_labeled)
    wb.save(out_path)
    print(f"  Saved: {out_path}")


def export_summary_excel(results, out_path):
    pivot = results.pivot_table(
        index="baseline", columns=["sut", "scenario"],
        values=["collision_rate_per100", "ttc_critical_rate_per100",
                "collision_count", "ttc_critical_count", "n_scenarios"],
        aggfunc="first",
    )
    known = [b for b in BASELINE_ORDER if b in pivot.index]
    pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
    pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
    pivot.to_excel(out_path, engine="openpyxl")

    wb = load_workbook(out_path)
    ws = wb.active
    hf = PatternFill("solid", fgColor="1F4E79")
    bw = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill, cell.font = hf, bw
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            max((len(str(c.value or "")) for c in col), default=10) + 4
        )
    wb.save(out_path)
    print(f"  Saved: {out_path}")


# ------------------------------------------------------------------------------
# FIGURES
# ------------------------------------------------------------------------------

def _draw_bars(ax, df_sut, metric, baselines, scenarios):
    n_s = len(scenarios)
    x   = np.arange(len(baselines))
    w   = 0.22
    off = np.linspace(-(n_s - 1) / 2, (n_s - 1) / 2, n_s) * w
    hatch = ["", "//", "xx"]
    alpha = [0.95, 0.80, 0.65]

    for si, scen in enumerate(scenarios):
        sub = df_sut[df_sut["scenario"] == scen].set_index("baseline")
        vals = [float(sub.loc[b, metric]) if b in sub.index else 0.0
                for b in baselines]
        bars = ax.bar(x + off[si], vals, width=w,
                      color=[BASELINE_COLORS.get(b, "#888") for b in baselines],
                      alpha=alpha[si], hatch=hatch[si],
                      edgecolor="white", linewidth=0.5,
                      label=f"Scenario {scen}", zorder=3)
        for bar, v in zip(bars, vals):
            if v > 0.3:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + 0.3, f"{v:.1f}",
                        ha="center", va="bottom", fontsize=6.5, color="#222",
                        clip_on=False)

    if "bayscen" in baselines:
        bx = baselines.index("bayscen")
        ax.axvspan(bx - 0.45, bx + 0.45, color="#E8F5E9", zorder=0, alpha=0.6)

    # Compute max value across all bars to set a tight ylim with headroom
    all_vals = []
    for scen in scenarios:
        sub = df_sut[df_sut["scenario"] == scen].set_index("baseline")
        all_vals += [float(sub.loc[b, metric]) if b in sub.index else 0.0 for b in baselines]
    max_val = max(all_vals) if all_vals else 1.0
    ax.set_ylim(bottom=0, top=max_val * 1.22)  # 22% headroom for labels

    ax.set_xticks(x)
    ax.yaxis.grid(True, linestyle="--", alpha=0.4, zorder=0)
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)
    return hatch, alpha


def plot_per_sut(results, metric, ylabel, out_path):
    scenarios = sorted(results["scenario"].unique())
    baselines = [b for b in BASELINE_ORDER if b in results["baseline"].values]

    fig, axes = plt.subplots(1, len(SUTS), figsize=(16, 5), sharey=False)
    for ax, sut in zip(axes, SUTS):
        hatch, alpha = _draw_bars(ax, results[results["sut"] == sut],
                                  metric, baselines, scenarios)
        ax.set_xticklabels([BASELINE_LABELS.get(b, b) for b in baselines],
                           rotation=30, ha="right", fontsize=8)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.set_title(sut, fontsize=10, fontweight="bold")
        if ax == axes[0]:
            patches = [Patch(facecolor="#AAAAAA", alpha=alpha[i],
                             hatch=hatch[i], label=f"Scenario {s}")
                       for i, s in enumerate(scenarios)]
            ax.legend(handles=patches, title="Scenario",
                      fontsize=7, title_fontsize=7, framealpha=0.85)

    fig.suptitle(f"RQ1 – {ylabel}", fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {out_path}")


def plot_combined(results, out_path):
    """2-row x 2-col: (collision | ttc) x (Interfuser | Modular)."""
    panels = [
        ("collision_rate_per100",   "Collision Rate (per 100 scenarios)"),
        ("ttc_critical_rate_per100","TTC-Critical Rate (per 100 scenarios)"),
    ]
    scenarios = sorted(results["scenario"].unique())
    baselines = [b for b in BASELINE_ORDER if b in results["baseline"].values]
    hatch = ["", "//", "xx"]
    alpha = [0.95, 0.80, 0.65]

    fig, axes = plt.subplots(len(panels), len(SUTS),
                             figsize=(16, 9), sharey="row")

    for ri, (metric, ylabel) in enumerate(panels):
        for ci, sut in enumerate(SUTS):
            ax = axes[ri][ci]
            _draw_bars(ax, results[results["sut"] == sut],
                       metric, baselines, scenarios)

            if ri == len(panels) - 1:
                ax.set_xticklabels(
                    [BASELINE_LABELS.get(b, b) for b in baselines],
                    rotation=30, ha="right", fontsize=8)
            else:
                ax.set_xticklabels([])
            if ci == 0:
                ax.set_ylabel(ylabel, fontsize=9)
            if ri == 0:
                ax.set_title(sut, fontsize=10, fontweight="bold", pad=6)
            if ri == 0 and ci == 0:
                patches = [Patch(facecolor="#AAAAAA", alpha=alpha[i],
                                 hatch=hatch[i], label=f"Scenario {s}")
                           for i, s in enumerate(scenarios)]
                ax.legend(handles=patches, title="Scenario",
                          fontsize=7, title_fontsize=7,
                          framealpha=0.85, loc="upper right")

    fig.suptitle("RQ1 – Criticality: Collision Rate and TTC-Critical Rate",
                 fontsize=12, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(out_path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {out_path}")


# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------

def main():
    os.makedirs(RESULTS_DIR, exist_ok=True)
    print("\n=== RQ1 – Criticality Analysis ===\n")
    print("Collecting results …")
    results = collect_all_results()
    print(f"\nLoaded {len(results)} groups  "
          f"({results['sut'].nunique()} SUTs, "
          f"{results['scenario'].nunique()} scenarios, "
          f"{results['baseline'].nunique()} baselines).\n")

    print("Exporting Excel …")
    export_raw_excel(results, os.path.join(RESULTS_DIR, "rq1_raw_data.xlsx"))
    export_summary_excel(results, os.path.join(RESULTS_DIR, "rq1_summary.xlsx"))

    print("\nGenerating figures …")
    plot_per_sut(results, "collision_rate_per100",
                 "Collision Rate (per 100 scenarios)",
                 os.path.join(RESULTS_DIR, "rq1_collision_rate.png"))
    plot_per_sut(results, "ttc_critical_rate_per100",
                 "TTC-Critical Rate (per 100 scenarios)",
                 os.path.join(RESULTS_DIR, "rq1_ttc_critical_rate.png"))
    plot_combined(results, os.path.join(RESULTS_DIR, "rq1_combined.png"))

    for metric, label in [("collision_rate_per100",   "Collision Rate (%)"),
                           ("ttc_critical_rate_per100","TTC-Critical Rate (%)")]:
        print(f"\n=== {label} ===")
        pivot = results.pivot_table(
            index="baseline", columns=["sut", "scenario"],
            values=metric, aggfunc="first")
        known = [b for b in BASELINE_ORDER if b in pivot.index]
        pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
        pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
        pivot.columns = [f"{s}/S{sc}" for s, sc in pivot.columns]
        print(pivot.round(2).to_string())

    print(f"\nAll RQ1 outputs saved to: {RESULTS_DIR}/\n")


if __name__ == "__main__":
    main()