"""
critical_plausibility.py
=============================
RQ2 supplement – Distribution of critical scenarios across physical violations.

For each method, identifies rows where:
  - algo_safety == 1  (safety-critical scenario)
  - at least one physical plausibility constraint is violated (C1–C6)

This answers: "Are the critical scenarios that BayScen finds physically
plausible, or are they artifacts of implausible parameter combinations?"

Constraints are identical to physical_plausibility.py:
  C1   P > 20 ⟹  D > 0               (meaningful precipitation causes deposits)
  C2   P > 20 ⟹  W > 0               (meaningful precipitation causes wetness)
  C3a  W < 40  ⟹  F ≤ 1 − W/200      (wetness reduces friction, low regime)
  C3b  W ≥ 40  ⟹  F ≤ 0.6            (wetness reduces friction, high regime)
  C4   |L − (100 − G)| ≤ ε=10        (fog density determines fog distance)
  C5   N ≥ 60  ⟹  G ≤ 40             (high wind disperses fog)
  C6   P > 20  ⟹  C > 0              (rain requires cloud cover)

METRICS COMPUTED (per method, aggregated across 3 runs)
-------------------------------------------------------
  n_critical          – total algo_safety=1 rows (mean across runs)
  n_critical_violated – critical rows with ≥1 constraint violated
  n_critical_clean    – critical rows with 0 violations (plausible & critical)
  violated_rate       – n_critical_violated / n_critical  (%)
  clean_rate          – n_critical_clean    / n_critical  (%)
  per-constraint counts among critical rows

OUTPUTS  →  ./results/rq2/
  critical_plausibility.xlsx
    - Sheet per SUT  (detailed per-run + mean)
    - Summary pivot  (clean_rate % by baseline × scenario)
    - Constraint breakdown among critical scenarios

USAGE
-----
  python scripts/critical_plausibility.py --results "simulation results"
"""

import argparse
import glob
import logging
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

SUTS = ["Interfuser", "Modular"]

SCENARIO_FOLDERS = {
    1: "Scenario 1",
    2: "Scenario 2",
    3: "Scenario 3",
}

BASELINE_ORDER = [
    "random", "sitcov", "pict_2way", "pict_3way",
    "ctbc", "avfuzzer", "bayscen_common", "bayscen",
]

BASELINE_LABELS = {
    "random":         "Random",
    "sitcov":         "SitCov",
    "pict_2way":      "PICT 2-way",
    "pict_3way":      "PICT 3-way",
    "ctbc":           "CTBC",
    "avfuzzer":       "AvFuzzer",
    "bayscen_common": "BayScen-Common",
    "bayscen":        "BayScen",
}

PREFIX_MAP = {
    "bayscen_common": "bayscen_common",
    "bayscen":        "bayscen",
    "avfuzzer":       "avfuzzer",
    "random":         "random",
    "sitcov":         "sitcov",
    "pict_2way":      "pict_2way",
    "pict_2w":        "pict_2way",
    "pict_3way":      "pict_3way",
    "pict_3w":        "pict_3way",
    "ctbc":           "ctbc",
}

# Constraint parameters
C4_EPSILON  = 10.0
P_THRESHOLD = 20.0

# Ordered constraint IDs — drives all loops in the file
CONSTRAINT_IDS = ["C1", "C2", "C3a", "C3b", "C4", "C5", "C6"]

# ──────────────────────────────────────────────────────────────────────────────
# Filename parsing
# ──────────────────────────────────────────────────────────────────────────────

def parse_filename(filename: str):
    stem = Path(filename).stem
    run_match = re.search(r"_run(\d+)$", stem, re.IGNORECASE)
    if not run_match:
        return None, None
    run_num = int(run_match.group(1))
    prefix_part = stem[: run_match.start()].lower()
    prefix_part = re.sub(r"_scenario\d+_\w+$", "", prefix_part)
    for key in sorted(PREFIX_MAP.keys(), key=len, reverse=True):
        if prefix_part == key or prefix_part.startswith(key):
            return PREFIX_MAP[key], run_num
    return prefix_part, run_num


# ──────────────────────────────────────────────────────────────────────────────
# Constraints  (identical to physical_plausibility.py)
# ──────────────────────────────────────────────────────────────────────────────

def _col(df: pd.DataFrame, *candidates: str) -> pd.Series:
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return df[lower_map[cand.lower()]]
    raise KeyError(f"None of {candidates} found in {df.columns.tolist()}")


def check_constraints(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return boolean DataFrame — True = violated.
    Identical logic to physical_plausibility.py.

    C1  : P > 20 AND D == 0
    C2  : P > 20 AND W == 0
    C3a : W < 40  AND F > 1 − W/200
    C3b : W >= 40 AND F > 0.6
    C4  : |L − (100 − G)| > ε
    C5  : N >= 60 AND G > 40
    C6  : P > 20  AND C == 0
    """
    P = _col(df, "feature_Precipitation")
    D = _col(df, "feature_PrecipitationDeposits")
    W = _col(df, "feature_Wetness")
    F = _col(df, "feature_RoadFriction")
    G = _col(df, "feature_FogDensity")
    L = _col(df, "feature_FogDistance")
    N = _col(df, "feature_WindIntensity")
    C = _col(df, "feature_Cloudiness")

    viol = pd.DataFrame(index=df.index)
    viol["C1"]  = (P > P_THRESHOLD) & (D == 0)
    viol["C2"]  = (P > P_THRESHOLD) & (W == 0)
    viol["C3a"] = (W < 40) & (F > (1.0 - W / 200.0))
    viol["C3b"] = (W >= 40) & (F > 0.6)
    viol["C4"]  = (L - (100 - G)).abs() > C4_EPSILON
    viol["C5"]  = (N >= 60) & (G > 40)
    viol["C6"]  = (P > P_THRESHOLD) & (C == 0)
    return viol


def get_algo_safety(df: pd.DataFrame) -> pd.Series:
    """Return boolean Series for algo_safety == 1, handling int/bool/float dtypes."""
    col = _col(df, "algo_safety")
    return col.astype(bool)


# ──────────────────────────────────────────────────────────────────────────────
# Per-run metric computation
# ──────────────────────────────────────────────────────────────────────────────

def compute_critical_plausibility(df: pd.DataFrame) -> dict:
    """
    For one run CSV:
      1. Flag critical rows (algo_safety == 1)
      2. Apply all constraints
      3. Among critical rows, count how many have ≥1 violation (implausible)
         and how many have 0 violations (plausible & critical)

    Returns a dict of counts and rates.
    """
    viol     = check_constraints(df)
    is_crit  = get_algo_safety(df)
    any_viol = viol.any(axis=1)

    n_total    = len(df)
    n_critical = int(is_crit.sum())

    crit_viol  = is_crit & any_viol
    crit_clean = is_crit & ~any_viol

    n_crit_violated = int(crit_viol.sum())
    n_crit_clean    = int(crit_clean.sum())

    violated_rate = n_crit_violated / n_critical * 100 if n_critical > 0 else 0.0
    clean_rate    = n_crit_clean    / n_critical * 100 if n_critical > 0 else 0.0

    # Per-constraint counts among critical rows
    per_c = {}
    for c in CONSTRAINT_IDS:
        cnt = int((is_crit & viol[c]).sum())
        per_c[f"{c}_critical_count"] = cnt
        per_c[f"{c}_critical_pct"]   = cnt / n_critical * 100 if n_critical > 0 else 0.0

    return {
        "n_total":           n_total,
        "n_critical":        n_critical,
        "critical_rate_pct": n_critical / n_total * 100 if n_total > 0 else 0.0,
        "n_crit_violated":   n_crit_violated,
        "n_crit_clean":      n_crit_clean,
        "violated_rate_pct": violated_rate,
        "clean_rate_pct":    clean_rate,
        **per_c,
    }


# ──────────────────────────────────────────────────────────────────────────────
# File indexing
# ──────────────────────────────────────────────────────────────────────────────

def index_files(results_dir: Path) -> dict:
    file_index: dict = {}
    for sut in SUTS:
        for scen_num, scen_folder in SCENARIO_FOLDERS.items():
            folder = results_dir / sut / scen_folder
            if not folder.is_dir():
                log.warning("Folder not found: %s", folder)
                continue
            for fpath in glob.glob(str(folder / "*.csv")):
                baseline, run_num = parse_filename(fpath)
                if baseline is None:
                    continue
                key = (sut, scen_num, baseline)
                file_index.setdefault(key, {})[run_num] = Path(fpath)
    return file_index


# ──────────────────────────────────────────────────────────────────────────────
# Main analysis loop
# ──────────────────────────────────────────────────────────────────────────────

def compute_all(file_index: dict) -> pd.DataFrame:
    records = []

    all_groups = [
        ((sut, scen, bl), runs)
        for (sut, scen, bl), runs in sorted(file_index.items())
        if len(runs) >= 3
    ]

    for (sut, scen, baseline), run_files in all_groups:
        label = BASELINE_LABELS.get(baseline, baseline)
        log.info("%-12s | Scen %d | %s", sut, scen, label)

        row = {"sut": sut, "scenario": scen, "baseline": baseline}
        per_run = []

        for r in (1, 2, 3):
            fpath = run_files.get(r)
            if fpath is None:
                log.warning("  Run %d missing", r)
                continue
            try:
                df = pd.read_csv(fpath)
                df.columns = df.columns.str.strip()
                metrics = compute_critical_plausibility(df)
                per_run.append(metrics)

                # Store per-run values
                row[f"n_critical_run{r}"]       = metrics["n_critical"]
                row[f"n_crit_clean_run{r}"]      = metrics["n_crit_clean"]
                row[f"n_crit_violated_run{r}"]   = metrics["n_crit_violated"]
                row[f"clean_rate_pct_run{r}"]    = metrics["clean_rate_pct"]
                row[f"violated_rate_pct_run{r}"] = metrics["violated_rate_pct"]

                log.info(
                    "  run%d  critical=%d  clean=%d (%.1f%%)  violated=%d (%.1f%%)",
                    r,
                    metrics["n_critical"],
                    metrics["n_crit_clean"],    metrics["clean_rate_pct"],
                    metrics["n_crit_violated"], metrics["violated_rate_pct"],
                )
            except Exception as exc:
                log.error("  run%d FAILED: %s", r, exc)

        if not per_run:
            continue

        # Aggregate across runs
        for key in ["n_critical", "n_crit_clean", "n_crit_violated",
                    "clean_rate_pct", "violated_rate_pct", "critical_rate_pct",
                    "n_total"]:
            vals = [m[key] for m in per_run]
            row[f"{key}_mean"] = float(np.mean(vals))
            row[f"{key}_std"]  = float(np.std(vals))

        for c in CONSTRAINT_IDS:
            pct_vals = [m[f"{c}_critical_pct"]   for m in per_run]
            cnt_vals = [m[f"{c}_critical_count"]  for m in per_run]
            row[f"{c}_crit_pct_mean"]   = float(np.mean(pct_vals))
            row[f"{c}_crit_pct_std"]    = float(np.std(pct_vals))
            row[f"{c}_crit_count_mean"] = float(np.mean(cnt_vals))

        row["n_scenarios"] = int(np.mean([m["n_total"] for m in per_run]))

        log.info(
            "  ✔  mean critical=%d  clean=%.1f%%  violated=%.1f%%",
            int(row["n_critical_mean"]),
            row["clean_rate_pct_mean"],
            row["violated_rate_pct_mean"],
        )
        records.append(row)

    if not records:
        raise RuntimeError("No data loaded. Check folder structure and filenames.")

    results = pd.DataFrame(records)
    order_map = {b: i for i, b in enumerate(BASELINE_ORDER)}
    results["_ord"] = results["baseline"].map(lambda b: order_map.get(b, 999))
    results = (results
               .sort_values(["sut", "scenario", "_ord"])
               .drop(columns="_ord")
               .reset_index(drop=True))
    return results


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def _style_sheet(ws, bayscen_rows: list):
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    bscn_fill = PatternFill("solid", fgColor="C8E6C9")
    alt_fill  = PatternFill("solid", fgColor="EEF2F7")
    bw   = Font(bold=True, color="FFFFFF")
    bb   = Font(bold=True)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDBDBD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, bw, ctr, brd
    ws.row_dimensions[1].height = 36

    for ri in range(2, ws.max_row + 1):
        is_bs = (ri - 1) in bayscen_rows
        for cell in ws[ri]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            if is_bs:
                cell.fill, cell.font = bscn_fill, bb
            elif ri % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 26)
    ws.freeze_panes = "C2"


def export_excel(results: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Detailed columns ───────────────────────────────────────────────────────
    detail_cols = (
        ["scenario", "baseline", "n_scenarios"]
        + [f"n_critical_run{r}"        for r in (1, 2, 3)]
        + [f"n_crit_clean_run{r}"      for r in (1, 2, 3)]
        + [f"clean_rate_pct_run{r}"    for r in (1, 2, 3)]
        + [f"violated_rate_pct_run{r}" for r in (1, 2, 3)]
        + ["n_critical_mean", "n_critical_std",
           "n_crit_clean_mean", "n_crit_violated_mean",
           "clean_rate_pct_mean", "clean_rate_pct_std",
           "violated_rate_pct_mean", "violated_rate_pct_std"]
        + [f"{c}_crit_pct_mean" for c in CONSTRAINT_IDS]
        + [f"{c}_crit_pct_std"  for c in CONSTRAINT_IDS]
    )

    header_map = {
        "scenario":               "Scenario",
        "baseline":               "Baseline",
        "n_scenarios":            "N Scenarios",
        "n_critical_mean":        "N Critical (mean)",
        "n_critical_std":         "N Critical (std)",
        "n_crit_clean_mean":      "N Clean Critical (mean)",
        "n_crit_violated_mean":   "N Violated Critical (mean)",
        "clean_rate_pct_mean":    "Clean Rate % (mean)",
        "clean_rate_pct_std":     "Clean Rate % (std)",
        "violated_rate_pct_mean": "Violated Rate % (mean)",
        "violated_rate_pct_std":  "Violated Rate % (std)",
        **{f"n_critical_run{r}":        f"N Critical run{r}"       for r in (1,2,3)},
        **{f"n_crit_clean_run{r}":      f"N Clean run{r}"          for r in (1,2,3)},
        **{f"clean_rate_pct_run{r}":    f"Clean % run{r}"          for r in (1,2,3)},
        **{f"violated_rate_pct_run{r}": f"Violated % run{r}"       for r in (1,2,3)},
        **{f"{c}_crit_pct_mean": f"{c} among critical % (mean)" for c in CONSTRAINT_IDS},
        **{f"{c}_crit_pct_std":  f"{c} among critical % (std)"  for c in CONSTRAINT_IDS},
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── One sheet per SUT ──────────────────────────────────────────────────
        for sut in SUTS:
            sub = results[results["sut"] == sut].copy()
            sub["baseline"] = sub["baseline"].map(lambda b: BASELINE_LABELS.get(b, b))
            existing = [c for c in detail_cols if c in sub.columns]
            out = sub[existing].copy()
            out.columns = [header_map.get(c, c) for c in existing]
            out.round(2).to_excel(writer, sheet_name=sut, index=False)

        # ── Summary pivot: clean_rate_pct_mean ────────────────────────────────
        pivot_clean = results.pivot_table(
            index="baseline", columns=["sut", "scenario"],
            values="clean_rate_pct_mean", aggfunc="first",
        )
        known = [b for b in BASELINE_ORDER if b in pivot_clean.index]
        pivot_clean = pivot_clean.reindex(
            known + [b for b in pivot_clean.index if b not in known]
        )
        pivot_clean.index = [BASELINE_LABELS.get(b, b) for b in pivot_clean.index]
        pivot_clean.columns = [f"{s}/S{sc}" for s, sc in pivot_clean.columns]
        pivot_clean.round(2).to_excel(
            writer, sheet_name="Clean Critical Rate (%) Summary"
        )

        # ── Summary pivot: n_critical_mean ────────────────────────────────────
        pivot_n = results.pivot_table(
            index="baseline", columns=["sut", "scenario"],
            values="n_critical_mean", aggfunc="first",
        )
        known = [b for b in BASELINE_ORDER if b in pivot_n.index]
        pivot_n = pivot_n.reindex(
            known + [b for b in pivot_n.index if b not in known]
        )
        pivot_n.index = [BASELINE_LABELS.get(b, b) for b in pivot_n.index]
        pivot_n.columns = [f"{s}/S{sc}" for s, sc in pivot_n.columns]
        pivot_n.round(1).to_excel(writer, sheet_name="N Critical Scenarios Summary")

        # ── Per-constraint breakdown among critical scenarios ──────────────────
        for c_id in CONSTRAINT_IDS:
            col = f"{c_id}_crit_pct_mean"
            if col not in results.columns:
                continue
            pivot_c = results.pivot_table(
                index="baseline", columns=["sut", "scenario"],
                values=col, aggfunc="first",
            )
            known_c = [b for b in BASELINE_ORDER if b in pivot_c.index]
            pivot_c = pivot_c.reindex(
                known_c + [b for b in pivot_c.index if b not in known_c]
            )
            pivot_c.index = [BASELINE_LABELS.get(b, b) for b in pivot_c.index]
            pivot_c.columns = [f"{s}/S{sc}" for s, sc in pivot_c.columns]
            pivot_c.round(2).to_excel(
                writer,
                sheet_name=f"{c_id} among critical (%)"[:31]
            )

    # ── Post-process styling ───────────────────────────────────────────────────
    wb = load_workbook(output_path)

    for sut in SUTS:
        if sut not in wb.sheetnames:
            continue
        ws = wb[sut]
        sub = results[results["sut"] == sut].reset_index(drop=True)
        bs_rows = [i + 1 for i, b in enumerate(sub["baseline"]) if b == "bayscen"]
        _style_sheet(ws, bs_rows)

    for sh in wb.sheetnames:
        if sh in SUTS:
            continue
        ws = wb[sh]
        hf = PatternFill("solid", fgColor="1F4E79")
        bw = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill, cell.font = hf, bw
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max((len(str(c.value or "")) for c in col), default=8) + 4
            )

    wb.save(output_path)
    log.info("Saved → %s", output_path)


# ──────────────────────────────────────────────────────────────────────────────
# Console summary
# ──────────────────────────────────────────────────────────────────────────────

def print_summary(results: pd.DataFrame) -> None:
    print(f"\n{'='*70}")
    print("  RQ2 Supplement – Critical Scenarios: Clean Rate %")
    print("  (% of safety-critical scenarios with NO physical violations)")
    print("  Higher = more plausible critical findings")
    print(f"{'='*70}")
    pivot = results.pivot_table(
        index="baseline", columns=["sut", "scenario"],
        values="clean_rate_pct_mean", aggfunc="first",
    )
    known = [b for b in BASELINE_ORDER if b in pivot.index]
    pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
    pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
    pivot.columns = [f"{s}/S{sc}" for s, sc in pivot.columns]
    print(pivot.round(2).to_string())

    print(f"\n{'='*70}")
    print("  N Critical Scenarios (mean across 3 runs)")
    print(f"{'='*70}")
    pivot2 = results.pivot_table(
        index="baseline", columns=["sut", "scenario"],
        values="n_critical_mean", aggfunc="first",
    )
    known2 = [b for b in BASELINE_ORDER if b in pivot2.index]
    pivot2 = pivot2.reindex(known2 + [b for b in pivot2.index if b not in known2])
    pivot2.index = [BASELINE_LABELS.get(b, b) for b in pivot2.index]
    pivot2.columns = [f"{s}/S{sc}" for s, sc in pivot2.columns]
    print(pivot2.round(1).to_string())

    print(f"\n{'='*70}")
    print("  Constraints active in this run")
    print(f"{'='*70}")
    print(f"  IDs : {', '.join(CONSTRAINT_IDS)}")
    print(f"  C4ε : {C4_EPSILON}   P threshold : {P_THRESHOLD}")


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="RQ2 supplement – Critical scenario plausibility (updated constraints).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--results", type=Path, default=Path("simulation results"),
        help="Root folder containing Interfuser/ and Modular/ sub-folders.",
    )
    p.add_argument(
        "--output", type=Path,
        default=Path("results/rq2/critical_plausibility.xlsx"),
        help="Output Excel file path.",
    )
    p.add_argument(
        "--epsilon", type=float, default=10.0,
        help="Tolerance ε for constraint C4: |L − (100−G)| ≤ ε.",
    )
    p.add_argument(
        "--precip-threshold", type=float, default=20.0,
        help="Precipitation threshold for C1, C2, C6 (default 20, i.e. P > 20).",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    global C4_EPSILON, P_THRESHOLD
    C4_EPSILON  = args.epsilon
    P_THRESHOLD = args.precip_threshold

    log.info("=" * 65)
    log.info("RQ2 Supplement – Critical Scenario Plausibility (updated constraints)")
    log.info("=" * 65)
    log.info("  Results dir      : %s", args.results)
    log.info("  Output           : %s", args.output)
    log.info("  C4 epsilon       : %.1f", C4_EPSILON)
    log.info("  Precip threshold : %.1f (C1, C2, C6)", P_THRESHOLD)
    log.info("  Constraints      : %s", ", ".join(CONSTRAINT_IDS))
    log.info("=" * 65)

    if not args.results.is_dir():
        log.error("Results directory not found: %s", args.results)
        raise SystemExit(1)

    log.info("\nIndexing CSV files …")
    file_index = index_files(args.results)
    n_groups = sum(1 for runs in file_index.values() if len(runs) >= 3)
    log.info("Found %d complete groups.\n", n_groups)

    if n_groups == 0:
        log.error("No complete groups found.")
        raise SystemExit(1)

    log.info("Computing critical plausibility …\n")
    results = compute_all(file_index)

    log.info("\nExporting Excel …")
    export_excel(results, args.output)

    print_summary(results)
    print(f"\nAll outputs saved to: {args.output}\n")


if __name__ == "__main__":
    main()
