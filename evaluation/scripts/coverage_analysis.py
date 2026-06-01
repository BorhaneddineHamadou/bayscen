"""
coverage_analysis.py
====================
RQ3 – Coverage Analysis for BayScen paper using TISA metrics.

Computes area_IS, area_bugs, and cov_IS for every method's generated
scenarios via MATLAB's buildIS.m (same pipeline as tisa.py / correlation_analysis.py).

FOLDER STRUCTURE EXPECTED
--------------------------
simulation results/
  Interfuser/
    Scenario 1/
      avfuzzer_scenario1_interfuser_run1.csv
      avfuzzer_scenario1_interfuser_run2.csv
      avfuzzer_scenario1_interfuser_run3.csv
      bayscen_scenario1_interfuser_run1.csv
      ...
    Scenario 2/ ...
    Scenario 3/ ...
  Modular/ ...

Each CSV must contain:
  feature_*    – environmental / scenario feature columns
  algo_safety  – 0 (safe) or 1 (bug / safety-critical)

tisa/ (sibling of scripts/)
  tisa.py
  tisa_matlab/
    ISA-code/
      InstanceSpace/   ← all .m files (buildIS.m, PILOT.m, …)
      ISA-coverages.csv

TISA METRIC MAPPING  (from ISA-coverages.csv → paper metric names)
  area_IS   ← Footprint_area
  area_bugs ← Good_footprint_area
  cov_IS    ← COV_prunnedBoundary   (MATLAB typo kept)

OUTPUT  →  ./results/rq3/tisa_coverage.xlsx
  One sheet per SUT (Interfuser / Modular), containing:
    - scenario, baseline
    - area_IS_run1/2/3, area_IS_mean, area_IS_std
    - area_bugs_run1/2/3, area_bugs_mean, area_bugs_std
    - cov_IS_run1/2/3, cov_IS_mean, cov_IS_std
    - n_scenarios, n_bugs, n_safe  (per run, averaged)

USAGE
-----
  # Windows
  python scripts/coverage_analysis.py ^
      --results  "simulation results" ^
      --tisa     tisa/tisa_matlab ^
      --output   results/rq3/tisa_coverage.xlsx ^
      --matlab   "C:\\Program Files\\MATLAB\\R2025b\\bin\\matlab.exe"

  # Linux / macOS
  python scripts/coverage_analysis.py \\
      --results  "simulation results" \\
      --tisa     tisa/tisa_matlab \\
      --output   results/rq3/tisa_coverage.xlsx \\
      --matlab   /usr/local/MATLAB/R2023b/bin/matlab
"""

import argparse
import json
import logging
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path

try:
    from tqdm import tqdm
except ImportError:
    import subprocess as _sp
    _sp.check_call([sys.executable, "-m", "pip", "install", "tqdm", "-q", "--break-system-packages"])
    from tqdm import tqdm

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# TISA / ISA options  (identical to tisa.py OPTS)
# ──────────────────────────────────────────────────────────────────────────────

TISA_OPTS = {
    "parallel":  {"flag": False, "ncores": 1},
    "perf":      {"MaxPerf": False, "AbsPerf": True,
                  "epsilon": 0.50, "betaThreshold": 0.55},
    "auto":      {"preproc": True},
    "bound":     {"flag": True},
    "norm":      {"flag": True},
    "sifted":    {"flag": False, "rho": 0.1, "K": 10,
                  "NTREES": 50, "MaxIter": 1000, "Replicates": 100},
    "pilot":     {"analytic": False, "ntries": 5},
    "cloister":  {"pval": 0.05, "cthres": 0.7},
    "trace":     {"usesim": False, "PI": 0.55},
    "selvars":   {"smallscaleflag": False, "smallscale": 0.50,
                  "fileidxflag": False, "fileidx": "",
                  "densityflag": False, "mindistance": 0.1},
    "outputs":   {"csv": True, "web": False, "png": True},
}

# autoNormalize.m patch for MATLAB R2019b+  (verbatim from tisa.py)
AUTONORMALIZE_PATCH = """\
function [X,Y,out] = autoNormalize(X, Y)
% Patched version: compatible with MATLAB R2019b+
nfeats = size(X,2);
nalgos = size(Y,2);
disp("-> Auto-normalizing the data using Box-Cox and Z transformations.");
out.minX = min(X,[],1);
X = bsxfun(@minus,X,out.minX)+1;
out.lambdaX = zeros(1,nfeats);
out.muX     = zeros(1,nfeats);
out.sigmaX  = zeros(1,nfeats);
for i=1:nfeats
    aux = X(:,i);
    idx = isnan(aux);
    lam = boxcox(aux(~idx));
    out.lambdaX(i) = lam;
    aux(~idx) = applyBoxCox(aux(~idx), lam);
    [aux, out.muX(i), out.sigmaX(i)] = zscore(aux);
    X(~idx,i) = aux(~idx);
end
out.minY = min(Y(:));
Y = (Y - out.minY) + eps;
out.lambdaY = zeros(1,nalgos);
out.muY     = zeros(1,nalgos);
out.sigmaY  = zeros(1,nalgos);
for i=1:nalgos
    aux = Y(:,i);
    idx = isnan(aux);
    lam = boxcox(aux(~idx));
    out.lambdaY(i) = lam;
    aux(~idx) = applyBoxCox(aux(~idx), lam);
    [aux, out.muY(i), out.sigmaY(i)] = zscore(aux);
    Y(~idx,i) = aux(~idx);
end
end

function y = applyBoxCox(x, lam)
if abs(lam) < 1e-8
    y = log(x);
else
    y = (x .^ lam - 1) ./ lam;
end
end
"""

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

SUTS = ["Interfuser", "Modular"]

SCENARIO_FOLDERS = {
    1: "Scenario 1",
    2: "Scenario 2",
    3: "Scenario 3",
}

# Canonical baseline display order and labels
BASELINE_ORDER = [
    "random", "sitcov", "pict_2way", "pict_3way",
    "ctbc", "avfuzzer", "bayscen_common", "bayscen",
]

BASELINE_LABELS = {
    "random":         "Random",
    "sitcov":         "SitCov",
    "pict_2way":      "PICT 2-way",
    "pict_3way":      "PICT 3-way",
    "ctbc":           "CTBC",
    "avfuzzer":       "AvFuzzer",
    "bayscen_common": "BayScen-Common",
    "bayscen":        "BayScen",
}

# Filename prefix → canonical key (longest match first avoids bayscen matching bayscen_common)
PREFIX_MAP = {
    "bayscen_common": "bayscen_common",
    "bayscen":        "bayscen",
    "avfuzzer":       "avfuzzer",
    "random":         "random",
    "sitcov":         "sitcov",
    "pict_2way":      "pict_2way",
    "pict_2w":        "pict_2way",
    "pict_3way":      "pict_3way",
    "pict_3w":        "pict_3way",
    "ctbc":           "ctbc",
}

NAN = float("nan")

# ──────────────────────────────────────────────────────────────────────────────
# Filename helpers  (reused from criticality.py logic)
# ──────────────────────────────────────────────────────────────────────────────

def parse_filename(filename: str):
    """
    Extract (baseline_key, run_number) from filenames such as:
      avfuzzer_scenario2_modular_run1.csv
      PICT_2w_scenario3_modular_run2.csv
      bayscen_common_scenario1_interfuser_run3.csv
    Returns (baseline_key, run_int) or (None, None).
    """
    stem = Path(filename).stem
    run_match = re.search(r"_run(\d+)$", stem, re.IGNORECASE)
    if not run_match:
        return None, None
    run_num = int(run_match.group(1))

    prefix_part = stem[: run_match.start()].lower()
    prefix_part = re.sub(r"_scenario\d+_\w+$", "", prefix_part)

    for key in sorted(PREFIX_MAP.keys(), key=len, reverse=True):
        if prefix_part == key or prefix_part.startswith(key):
            return PREFIX_MAP[key], run_num

    return prefix_part, run_num   # unknown baseline – keep raw name


# ──────────────────────────────────────────────────────────────────────────────
# Data loading
# ──────────────────────────────────────────────────────────────────────────────

def load_and_prepare_csv(filepath: Path) -> pd.DataFrame:
    """
    Load a per-run CSV, validate TISA-required columns, encode
    categorical feature_* columns to integers (MATLAB requirement).
    """
    df = pd.read_csv(filepath)
    df.columns = df.columns.str.strip()

    # Normalise column names
    rename = {}
    for c in df.columns:
        cl = c.lower()
        if cl == "algo_safety":
            rename[c] = "algo_safety"
        elif cl.startswith("feature_"):
            rename[c] = cl
        else:
            rename[c] = cl
    df = df.rename(columns=rename)

    if "algo_safety" not in df.columns:
        raise ValueError(f"'algo_safety' column not found in {filepath}")
    feature_cols = [c for c in df.columns if c.startswith("feature_")]
    if not feature_cols:
        raise ValueError(f"No 'feature_*' columns found in {filepath}")

    # Label-encode categorical features for MATLAB
    for col in feature_cols:
        if df[col].dtype == object or str(df[col].dtype) == "category":
            uv = sorted(df[col].dropna().unique(), key=str)
            df[col] = df[col].map({v: i for i, v in enumerate(uv)})

    return df


# ──────────────────────────────────────────────────────────────────────────────
# TISA / MATLAB pipeline  (ported from tisa.py + correlation_analysis.py)
# ──────────────────────────────────────────────────────────────────────────────

def _norm_path(s) -> str:
    return str(s).replace("\\", "/").rstrip("/")


def _patch_autonormalize(matlab_code_dir: Path) -> None:
    target = matlab_code_dir / "autoNormalize.m"
    backup = matlab_code_dir / "autoNormalize.m.bak"
    if not target.exists():
        log.warning("autoNormalize.m not found at %s – skipping patch", target)
        return
    if not backup.exists():
        shutil.copy(target, backup)
        log.info("Backed up autoNormalize.m → %s", backup)
    target.write_text(AUTONORMALIZE_PATCH, encoding="utf-8")


def _write_options(run_dir: Path) -> None:
    op = run_dir / "options.json"
    if op.exists():
        op.unlink()
    op.write_text(json.dumps(TISA_OPTS, indent=2), encoding="utf-8")


def _write_launcher(run_dir: Path, matlab_code_dir: Path, cov_dir: Path) -> Path:
    launcher = run_dir / "run_buildIS.m"
    launcher.write_text(
        f"addpath('{matlab_code_dir.resolve().as_posix()}');\n"
        f"try\n"
        f"    buildIS('{run_dir.resolve().as_posix()}/',"
        f"'{cov_dir.resolve().as_posix()}/');\n"
        f"    disp('EOF:SUCCESS');\n"
        f"catch ME\n"
        f"    disp('EOF:ERROR'); disp(ME.message);\n"
        f"end\nexit;\n",
        encoding="utf-8",
    )
    return launcher


def _run_matlab(matlab_bin: str, launcher: Path, run_dir: Path,
                timeout: int = 1800) -> bool:
    is_win = sys.platform.startswith("win")
    if is_win:
        cmd = [matlab_bin, "-nosplash", "-nodesktop", "-wait",
               "-batch", launcher.stem]
        cwd = str(run_dir.resolve())
    else:
        cmd = [matlab_bin, "-nodisplay", "-nosplash", "-nodesktop",
               "-r", f"run('{launcher.as_posix()}')"]
        cwd = None

    try:
        res = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True,
            encoding="utf-8", errors="replace", timeout=timeout,
        )
        output = res.stdout + res.stderr
        (run_dir / "matlab_run.log").write_text(output, encoding="utf-8")
        return "EOF:SUCCESS" in output
    except subprocess.TimeoutExpired:
        log.warning("MATLAB timeout after %ds", timeout)
        return False
    except Exception as exc:
        log.warning("MATLAB subprocess error: %s", exc)
        return False


def _extract_metrics(run_dir: Path, coverages_csv: Path) -> dict:
    """Read ISA-coverages.csv and extract the three TISA metrics for this run."""
    nan_r = {"area_IS": NAN, "area_bugs": NAN, "cov_IS": NAN}
    if not coverages_csv.exists():
        log.warning("ISA-coverages.csv not found: %s", coverages_csv)
        return nan_r
    try:
        cov_df = pd.read_csv(coverages_csv)
        target = _norm_path(run_dir.resolve())
        match  = cov_df[cov_df["Rootdir"].apply(_norm_path) == target]
        row    = match.iloc[-1] if not match.empty else cov_df.iloc[-1]
        if match.empty:
            log.warning("  Rootdir not matched – using last row of ISA-coverages.csv")
        return {
            "area_IS":   float(row["Footprint_area"]),
            "area_bugs": float(row["Good_footprint_area"]),
            "cov_IS":    float(row["COV_prunnedBoundary"]),
        }
    except Exception as exc:
        log.warning("  Metric extraction failed: %s", exc)
        return nan_r


def run_tisa_for_csv(
    df: pd.DataFrame,
    run_dir: Path,
    tisa_root: Path,
    matlab_bin: str,
) -> dict:
    """
    Full TISA pipeline for a single CSV (one run of one method):
      1. Write metadata.csv + options.json into a temp run_dir
      2. Patch autoNormalize.m
      3. Write + run MATLAB launcher
      4. Extract metrics from ISA-coverages.csv
    Returns dict with area_IS, area_bugs, cov_IS (NaN on failure).
    """
    nan_r = {"area_IS": NAN, "area_bugs": NAN, "cov_IS": NAN}

    matlab_code_dir = tisa_root / "ISA-code" / "InstanceSpace"
    cov_dir         = tisa_root / "ISA-code"
    coverages_csv   = cov_dir / "ISA-coverages.csv"

    if not matlab_code_dir.exists():
        log.error("InstanceSpace not found: %s", matlab_code_dir)
        return nan_r

    run_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(run_dir / "metadata.csv", index=False)
    _write_options(run_dir)
    _patch_autonormalize(matlab_code_dir)
    launcher = _write_launcher(run_dir, matlab_code_dir, cov_dir)

    t0 = time.perf_counter()
    success = _run_matlab(matlab_bin, launcher, run_dir)
    elapsed = time.perf_counter() - t0

    if not success:
        log.warning("  MATLAB did not succeed (%.1f s) – check %s",
                    elapsed, run_dir / "matlab_run.log")
        return nan_r

    log.info("  MATLAB OK in %.1f s", elapsed)
    return _extract_metrics(run_dir, coverages_csv)


# ──────────────────────────────────────────────────────────────────────────────
# File indexing  (same as criticality.py)
# ──────────────────────────────────────────────────────────────────────────────

def index_files(results_dir: Path) -> dict:
    """
    Walk results_dir and return:
      { (sut, scen_num, baseline): {run_num: Path} }
    """
    import glob
    file_index: dict = {}

    for sut in SUTS:
        for scen_num, scen_folder in SCENARIO_FOLDERS.items():
            folder = results_dir / sut / scen_folder
            if not folder.is_dir():
                log.warning("Folder not found: %s", folder)
                continue
            for fpath in glob.glob(str(folder / "*.csv")):
                baseline, run_num = parse_filename(fpath)
                if baseline is None:
                    log.warning("Cannot parse filename: %s", fpath)
                    continue
                key = (sut, scen_num, baseline)
                file_index.setdefault(key, {})[run_num] = Path(fpath)

    return file_index


# ──────────────────────────────────────────────────────────────────────────────
# Main evaluation loop
# ──────────────────────────────────────────────────────────────────────────────

def compute_all_tisa(
    file_index: dict,
    tisa_root: Path,
    matlab_bin: str,
    workspace: Path,
) -> pd.DataFrame:
    """
    For each (sut, scenario, baseline) group:
      - Load and prepare run1, run2, run3
      - Run TISA on each run (sequentially – shared ISA-coverages.csv)
      - Record per-run metrics + mean + std

    Returns a long-format DataFrame with one row per group.
    """
    records = []
    all_groups = [
        ((sut, scen_num, baseline), run_files)
        for (sut, scen_num, baseline), run_files in sorted(file_index.items())
        if len(sorted(run_files.keys())) >= 3
    ]
    # Also log skipped groups
    for (sut, scen_num, baseline), run_files in sorted(file_index.items()):
        if len(sorted(run_files.keys())) < 3:
            log.warning("[SKIP] %s | Scen %d | %s – only %d run(s) found %s",
                        sut, scen_num, baseline,
                        len(run_files), sorted(run_files.keys()))

    total_runs = len(all_groups) * 3   # 3 MATLAB calls per group

    # ── Outer bar: groups (sut × scenario × baseline) ─────────────────────────
    group_bar = tqdm(
        all_groups,
        desc="Groups ",
        unit="group",
        colour="green",
        dynamic_ncols=True,
        position=0,
        leave=True,
    )
    # ── Inner bar: individual MATLAB runs (run1/2/3) ───────────────────────────
    run_bar = tqdm(
        total=total_runs,
        desc="MATLAB ",
        unit="run",
        colour="cyan",
        dynamic_ncols=True,
        position=1,
        leave=True,
    )

    done = 0
    for (sut, scen_num, baseline), run_files in group_bar:
        done += 1
        label = BASELINE_LABELS.get(baseline, baseline)
        group_bar.set_description(
            f"Groups  [{sut[:3]} S{scen_num} {label[:14]}]"
        )
        log.info("")
        log.info("━" * 65)
        log.info("[%d/%d]  %s  |  Scenario %d  |  %s",
                 done, len(all_groups), sut, scen_num, baseline)
        log.info("━" * 65)

        row = {
            "sut":      sut,
            "scenario": scen_num,
            "baseline": baseline,
        }

        # Per-run TISA
        for metric in ("area_IS", "area_bugs", "cov_IS"):
            for r in (1, 2, 3):
                row[f"{metric}_run{r}"] = NAN

        run_counts = {"n_scenarios": [], "n_bugs": [], "n_safe": []}

        for r in (1, 2, 3):
            fpath = run_files.get(r)
            if fpath is None:
                log.warning("  Run %d missing for %s | Scen %d | %s",
                            r, sut, scen_num, baseline)
                run_bar.update(1)
                continue

            run_bar.set_description(
                f"MATLAB  [{sut[:3]} S{scen_num} {label[:10]} run{r}]"
            )
            log.info("  ── Run %d  →  %s", r, fpath.name)

            try:
                df = load_and_prepare_csv(fpath)
            except Exception as exc:
                log.error("  Load failed: %s", exc)
                run_bar.update(1)
                continue

            n_bugs = int((df["algo_safety"] == 1).sum())
            n_safe = int((df["algo_safety"] == 0).sum())
            run_counts["n_scenarios"].append(len(df))
            run_counts["n_bugs"].append(n_bugs)
            run_counts["n_safe"].append(n_safe)

            log.info("    n=%d  bugs=%d  safe=%d", len(df), n_bugs, n_safe)

            # Unique run dir inside workspace to avoid MATLAB collisions
            run_label = f"{sut}_{scen_num}_{baseline}_run{r}".replace(" ", "_")
            run_dir   = workspace / run_label

            metrics = run_tisa_for_csv(df, run_dir, tisa_root, matlab_bin)

            for metric, val in metrics.items():
                row[f"{metric}_run{r}"] = val
                log.info("    %s = %.4f", metric, val if not np.isnan(val) else -1)

            run_bar.update(1)

        # Aggregate across runs
        for metric in ("area_IS", "area_bugs", "cov_IS"):
            vals = [row[f"{metric}_run{r}"] for r in (1, 2, 3)
                    if not np.isnan(row[f"{metric}_run{r}"])]
            row[f"{metric}_mean"] = float(np.mean(vals))  if vals else NAN
            row[f"{metric}_std"]  = float(np.std(vals))   if len(vals) > 1 else NAN
            row[f"{metric}_min"]  = float(np.min(vals))   if vals else NAN
            row[f"{metric}_max"]  = float(np.max(vals))   if vals else NAN

        # Dataset stats (mean across runs)
        for key, vals in run_counts.items():
            row[key] = int(np.mean(vals)) if vals else 0

        records.append(row)
        group_bar.write(
            f"  ✔  [{sut} S{scen_num} {label}] "
            f"area_IS={row.get('area_IS_mean', NAN):.4f} | "
            f"area_bugs={row.get('area_bugs_mean', NAN):.4f} | "
            f"cov_IS={row.get('cov_IS_mean', NAN):.2f}%"
        )

    group_bar.close()
    run_bar.close()

    if not records:
        raise RuntimeError("No results computed. Check folder structure, filenames, and MATLAB.")

    results = pd.DataFrame(records)

    # Sort in canonical baseline order
    order_map = {b: i for i, b in enumerate(BASELINE_ORDER)}
    results["_ord"] = results["baseline"].map(lambda b: order_map.get(b, 999))
    results = (results.sort_values(["sut", "scenario", "_ord"])
                      .drop(columns="_ord")
                      .reset_index(drop=True))
    return results


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def _style_sheet(ws, n_data_rows: int, bayscen_rows: list[int]):
    """Apply header + alternating + BayScen highlight styling."""
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    bscn_fill = PatternFill("solid", fgColor="C8E6C9")
    alt_fill  = PatternFill("solid", fgColor="EEF2F7")
    bw   = Font(bold=True, color="FFFFFF")
    bb   = Font(bold=True)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDBDBD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, bw, ctr, brd
    ws.row_dimensions[1].height = 30

    # Data rows
    for ri in range(2, n_data_rows + 2):
        is_bs = (ri - 1) in bayscen_rows   # 1-indexed data rows
        for cell in ws[ri]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            if is_bs:
                cell.fill, cell.font = bscn_fill, bb
            elif ri % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 22)

    ws.freeze_panes = "C2"


def export_excel(results: pd.DataFrame, output_path: Path) -> None:
    """
    Write tisa_coverage.xlsx with:
      - One sheet per SUT (Interfuser / Modular)
      - Columns: scenario | baseline | per-run metrics | mean | std | min | max | dataset stats
      - A Summary sheet with mean values pivoted for quick comparison
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build display columns
    metric_cols_ordered = []
    for metric in ("area_IS", "area_bugs", "cov_IS"):
        metric_cols_ordered += [
            f"{metric}_run1", f"{metric}_run2", f"{metric}_run3",
            f"{metric}_mean", f"{metric}_std",
        ]

    display_cols = (
        ["scenario", "baseline"]
        + metric_cols_ordered
        + ["n_scenarios", "n_bugs", "n_safe"]
    )

    # Friendly headers
    header_map = {
        "scenario":          "Scenario",
        "baseline":          "Baseline",
        "n_scenarios":       "N Scenarios",
        "n_bugs":            "N Bugs",
        "n_safe":            "N Safe",
    }
    for metric, label in [("area_IS",   "area_IS"),
                           ("area_bugs", "area_bugs"),
                           ("cov_IS",    "cov_IS (%)")]:
        for r in (1, 2, 3):
            header_map[f"{metric}_run{r}"] = f"{label} run{r}"
        header_map[f"{metric}_mean"] = f"{label} mean"
        header_map[f"{metric}_std"]  = f"{label} std"

    display_labels = [header_map.get(c, c) for c in display_cols]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── One sheet per SUT ─────────────────────────────────────────────────
        for sut in SUTS:
            sub = results[results["sut"] == sut].copy()
            sub["baseline"] = sub["baseline"].map(lambda b: BASELINE_LABELS.get(b, b))

            # Keep only columns that exist in sub
            existing_cols = [c for c in display_cols if c in sub.columns]
            existing_labels = [header_map.get(c, c) for c in existing_cols]

            out = sub[existing_cols].copy()
            out.columns = existing_labels
            out.to_excel(writer, sheet_name=sut, index=False)

        # ── Summary pivot sheet ───────────────────────────────────────────────
        for metric, label in [("area_IS_mean",   "area_IS mean"),
                               ("area_bugs_mean", "area_bugs mean"),
                               ("cov_IS_mean",    "cov_IS mean (%)")]:
            if metric not in results.columns:
                continue
            pivot = results.pivot_table(
                index="baseline", columns=["sut", "scenario"],
                values=metric, aggfunc="first",
            )
            known = [b for b in BASELINE_ORDER if b in pivot.index]
            pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
            pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
            pivot.columns = [f"{s}/S{sc}" for s, sc in pivot.columns]
            pivot.round(4).to_excel(writer, sheet_name=label[:31])  # Excel 31-char limit

    # ── Post-process: apply styling ───────────────────────────────────────────
    wb = load_workbook(output_path)

    for sut in SUTS:
        if sut not in wb.sheetnames:
            continue
        ws = wb[sut]
        sub = results[results["sut"] == sut].reset_index(drop=True)
        # Identify BayScen rows (1-indexed in sheet = data_row_index + 1, header=row1)
        bayscen_rows = [i + 1 for i, b in enumerate(sub["baseline"]) if b == "bayscen"]
        _style_sheet(ws, len(sub), bayscen_rows)

    # Style summary sheets too (light only)
    for sh in wb.sheetnames:
        if sh in SUTS:
            continue
        ws = wb[sh]
        hf = PatternFill("solid", fgColor="1F4E79")
        bw = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill, cell.font = hf, bw
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max((len(str(c.value or "")) for c in col), default=8) + 4
            )

    wb.save(output_path)
    log.info("Saved → %s", output_path)


# ──────────────────────────────────────────────────────────────────────────────
# Console summary
# ──────────────────────────────────────────────────────────────────────────────

def print_summary(results: pd.DataFrame) -> None:
    for metric, label in [("area_IS_mean",   "area_IS   (test suite diversity)"),
                           ("area_bugs_mean", "area_bugs (failure diversity)"),
                           ("cov_IS_mean",    "cov_IS    (ODD coverage %)")]:
        if metric not in results.columns:
            continue
        print(f"\n{'='*70}")
        print(f"  {label}")
        print(f"{'='*70}")
        pivot = results.pivot_table(
            index="baseline", columns=["sut", "scenario"],
            values=metric, aggfunc="first",
        )
        known = [b for b in BASELINE_ORDER if b in pivot.index]
        pivot = pivot.reindex(known + [b for b in pivot.index if b not in known])
        pivot.index = [BASELINE_LABELS.get(b, b) for b in pivot.index]
        pivot.columns = [f"{s}/S{sc}" for s, sc in pivot.columns]
        print(pivot.round(4).to_string())


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="RQ3 – TISA coverage metrics for BayScen evaluation.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--results", type=Path, default=Path("simulation results"),
        help="Root folder containing Interfuser/ and Modular/ sub-folders.",
    )
    p.add_argument(
        "--tisa", type=Path, required=True,
        help=(
            "Path to tisa_matlab/ folder, i.e. the folder containing "
            "ISA-code/InstanceSpace/ and ISA-code/ISA-coverages.csv."
        ),
    )
    p.add_argument(
        "--output", type=Path, default=Path("results/rq3/tisa_coverage.xlsx"),
        help="Output Excel file.",
    )
    p.add_argument(
        "--matlab",
        type=str,
        default=shutil.which("matlab") or "matlab",
        help=(
            "Full path to the MATLAB executable. "
            r'Windows: "C:\Program Files\MATLAB\R2025b\bin\matlab.exe"  '
            'Linux: /usr/local/MATLAB/R2023b/bin/matlab'
        ),
    )
    p.add_argument(
        "--workspace", type=Path, default=Path("results/rq3/tisa_runs"),
        help=(
            "Temporary folder where per-run TISA datasets are written. "
            "Each run gets its own sub-folder to avoid collisions."
        ),
    )
    return p.parse_args()


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    args = parse_args()

    log.info("=" * 65)
    log.info("RQ3 – Coverage Analysis (TISA Metrics)")
    log.info("=" * 65)
    log.info("  Results dir : %s", args.results)
    log.info("  TISA root   : %s", args.tisa)
    log.info("  MATLAB      : %s", args.matlab)
    log.info("  Output      : %s", args.output)
    log.info("  Workspace   : %s", args.workspace)
    log.info("=" * 65)

    # Guards
    matlab_code_dir = args.tisa / "ISA-code" / "InstanceSpace"
    if not matlab_code_dir.exists():
        log.error(
            "InstanceSpace not found: %s\n"
            "Check --tisa points to the tisa_matlab folder.", matlab_code_dir
        )
        sys.exit(1)

    if not shutil.which(args.matlab) and not Path(args.matlab).exists():
        log.error("MATLAB not found at '%s'. Set --matlab to the full path.", args.matlab)
        sys.exit(1)

    if not args.results.is_dir():
        log.error("Results directory not found: %s", args.results)
        sys.exit(1)

    args.workspace.mkdir(parents=True, exist_ok=True)

    # Index all CSV files
    log.info("\nIndexing CSV files …")
    file_index = index_files(args.results)
    n_groups = sum(1 for runs in file_index.values() if len(runs) >= 3)
    log.info("Found %d complete groups (≥3 runs) across %d entries.",
             n_groups, len(file_index))

    if n_groups == 0:
        log.error("No complete groups found. Check folder structure and filenames.")
        sys.exit(1)

    # Compute TISA for every run of every group
    log.info("\nRunning TISA via MATLAB (sequential – shared ISA-coverages.csv) …")
    log.info("Expected time: ~2-5 min per run × %d runs = several hours.", n_groups * 3)

    results = compute_all_tisa(
        file_index  = file_index,
        tisa_root   = args.tisa,
        matlab_bin  = args.matlab,
        workspace   = args.workspace,
    )

    log.info("\nExporting Excel …")
    export_excel(results, args.output)

    print_summary(results)

    print(f"\nAll RQ3 outputs saved to: {args.output}\n")


if __name__ == "__main__":
    main()