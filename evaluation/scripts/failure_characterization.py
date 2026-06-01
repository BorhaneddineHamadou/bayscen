"""
failure_characterization.py
============================
RQ4 – Failure Characterization for BayScen paper.

For each abstract capability variable (Sensor_Perception, Surface_Traction,
Lateral_Stability, Conflict_Geometry), computes the failure rate at each
degradation level, pooled across the 3 runs of BayScen and BayScen-Common,
for every SUT × scenario combination.

ABSTRACT VARIABLE DEFINITIONS  (ISO 34503 / paper §II.C)
----------------------------------------------------------
  Sensor_Perception        (a_perc)  – sensor perception degradation
    Parents: FogDensity, FogDistance(*), Cloudiness, Precipitation, TimeOfDay(*)
    * FogDistance : INVERTED  (0 = immediate dense fog = worst)
    * TimeOfDay   : INVERTED  (-90° night = worst, 90° noon = best)

  Surface_Traction       (a_trac)  – surface traction degradation
    Parents: RoadFriction(*), Wetness, PrecipitationDeposits
    * RoadFriction : INVERTED  (0.1 icy = worst, 1.0 dry = best)

  Lateral_Stability  (a_stab)  – lateral stability degradation
    Parent: WindIntensity  (0 = calm = best, 100 = storm = worst)

  Conflict_Geometry   (g)      – conflict geometry (scenarios 1 & 2 only)
    Source: the 'Conflict_Geometry' column directly (c1 / c2 / c4)

Each parent is mapped to a degradation level 0..5 using rank-based
normalisation (0 = least degraded, 5 = most degraded). The abstract
variable value is the mean of its parents' degradation levels, then
binned into 6 equal-width bins [0,1), [1,2), …, [5,5] → labels 0..5.

FILE STRUCTURE
--------------
simulation results/
  Interfuser/
    Scenario 1/   bayscen_scenario1_interfuser_run1.csv  …run2  …run3
                  bayscen_common_scenario1_interfuser_run1.csv  …
    Scenario 2/   …
    Scenario 3/   …   (no Conflict_Geometry column here)
  Modular/        …

Each CSV must contain:
  feature_FogDensity, feature_FogDistance, feature_Cloudiness,
  feature_Precipitation, feature_TimeOfDay (or feature_TimeofDay),
  feature_RoadFriction, feature_Wetness, feature_PrecipitationDeposits,
  feature_WindIntensity,
  Conflict_Geometry  (scenarios 1 & 2 only),
  algo_safety      (0 = safe, 1 = failure)

OUTPUT  →  ./results/rq4/failure_characterization.xlsx
  Sheets:
    RawData          – one row per (sut, scenario, baseline, run, abstract_var, level)
                       with n_scenarios, n_failures, failure_rate
    FailureRate_Mean – pivot: mean failure rate across runs,
                       rows=(sut, scenario, baseline, abstract_var),
                       columns=level_0 … level_5
    Summary          – wide table combining all abstract vars for quick comparison
    Parameters       – run configuration

USAGE
-----
  # Windows
  python scripts/failure_characterization.py ^
      --results  "simulation results" ^
      --output   results/rq4/failure_characterization.xlsx

  # Linux / macOS
  python scripts/failure_characterization.py \\
      --results  "simulation results" \\
      --output   results/rq4/failure_characterization.xlsx

  # Restrict to specific SUTs or scenarios
  python scripts/failure_characterization.py \\
      --results  "simulation results" \\
      --output   results/rq4/failure_characterization.xlsx \\
      --suts     Interfuser Modular \\
      --scenarios 1 2 3
"""

import argparse
import glob
import logging
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

TTC_THRESHOLD = 0.5
TTC_EPSILON   = 0.001   # absorbs CSV rounding artefacts (0.4999 → 0.500)
TTC_SENTINEL  = 9999.0

SUTS_DEFAULT      = ["Interfuser", "Modular"]
SCENARIOS_DEFAULT = [1, 2, 3]
SCENARIO_FOLDERS  = {1: "Scenario 1", 2: "Scenario 2", 3: "Scenario 3"}

# Only BayScen and BayScen-Common for RQ4
TARGET_BASELINES = {
    "bayscen_common": "BayScen-Common",
    "bayscen":        "BayScen",
}

# Conflict_Geometry is absent in Scenario 3
ABSTRACT_VARS_ALL   = ["Sensor_Perception", "Surface_Traction", "Lateral_Stability", "Conflict_Geometry"]
ABSTRACT_VARS_S3    = ["Sensor_Perception", "Surface_Traction", "Lateral_Stability"]   # no Conflict_Geometry

N_LEVELS = 6   # 0 .. 5

# ──────────────────────────────────────────────────────────────────────────────
# Abstract variable computation
# ──────────────────────────────────────────────────────────────────────────────

def _rank_to_level(series: pd.Series, invert: bool = False) -> pd.Series:
    """
    Map a numeric Series to integer degradation levels 0..5 using
    rank-based ordinal encoding.

    invert=True  → higher raw value = LESS degraded (e.g. RoadFriction 1.0 = best)
                   so we flip before ranking so 0 = least degraded output.
    invert=False → higher raw value = MORE degraded (direct mapping).

    The mapping is derived from the sorted unique values in *this specific
    Series*, so it is robust to missing levels in a single run file.
    Falls back to min-max normalisation scaled to 0..5 if more than 6
    unique values are present (e.g. TimeOfDay has 7 values: –90..90).
    """
    unique_vals = sorted(series.dropna().unique(), reverse=invert)
    if len(unique_vals) <= N_LEVELS:
        val_to_level = {v: i for i, v in enumerate(unique_vals)}
        return series.map(val_to_level).astype(float)
    else:
        # Min-max → 0..5 continuous, then round to int
        lo, hi = series.min(), series.max()
        if invert:
            normed = (hi - series) / (hi - lo) if hi != lo else pd.Series(0.0, index=series.index)
        else:
            normed = (series - lo) / (hi - lo) if hi != lo else pd.Series(0.0, index=series.index)
        return (normed * (N_LEVELS - 1)).round().astype(float)


def compute_abstract_variables(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add abstract capability columns to df:
      Sensor_Perception        – mean degradation level of perception parents (0-5)
      Surface_Traction       – mean degradation level of traction parents (0-5)
      Lateral_Stability  – degradation level of wind intensity (0-5)
      Conflict_Geometry   – raw value from Conflict_Geometry column (c1/c2/c4)
                          or NaN if absent (Scenario 3)

    All output values are in [0, 5].  Higher = more degraded.
    """
    df = df.copy()

    # ── Helper: find column case-insensitively ────────────────────────────────
    col_map = {c.lower(): c for c in df.columns}

    def get(name: str) -> pd.Series | None:
        return df[col_map[name.lower()]] if name.lower() in col_map else None

    # ── Sensor_Perception (a_perc) ────────────────────────────────────────────────────
    # Parents: FogDensity(+), FogDistance(–), Cloudiness(+), Precipitation(+), TimeOfDay(–)
    vis_parts = []
    fog_d = get("feature_FogDensity")
    fog_f = get("feature_FogDistance")
    cloud = get("feature_Cloudiness")
    precip= get("feature_Precipitation")
    tod   = get("feature_TimeOfDay"); tod = get("feature_TimeofDay") if tod is None else tod

    if fog_d  is not None: vis_parts.append(_rank_to_level(fog_d,  invert=False))
    if fog_f  is not None: vis_parts.append(_rank_to_level(fog_f,  invert=True))   # 0=dense=worst
    if cloud  is not None: vis_parts.append(_rank_to_level(cloud,  invert=False))
    if precip is not None: vis_parts.append(_rank_to_level(precip, invert=False))
    if tod    is not None: vis_parts.append(_rank_to_level(tod,    invert=True))    # night=worst

    df["Sensor_Perception"] = (
        pd.concat(vis_parts, axis=1).mean(axis=1) if vis_parts
        else pd.Series(np.nan, index=df.index)
    )

    # ── Surface_Traction (a_trac) ───────────────────────────────────────────────────
    # Parents: RoadFriction(–), Wetness(+), PrecipitationDeposits(+)
    rsurf_parts = []
    fric = get("feature_RoadFriction")
    wet  = get("feature_Wetness")
    dep  = get("feature_PrecipitationDeposits")

    if fric is not None: rsurf_parts.append(_rank_to_level(fric, invert=True))   # 0.1=icy=worst
    if wet  is not None: rsurf_parts.append(_rank_to_level(wet,  invert=False))
    if dep  is not None: rsurf_parts.append(_rank_to_level(dep,  invert=False))

    df["Surface_Traction"] = (
        pd.concat(rsurf_parts, axis=1).mean(axis=1) if rsurf_parts
        else pd.Series(np.nan, index=df.index)
    )

    # ── Lateral_Stability (a_stab) ──────────────────────────────────────────────
    # Parent: WindIntensity(+)
    wind = get("feature_WindIntensity")
    df["Lateral_Stability"] = (
        _rank_to_level(wind, invert=False) if wind is not None
        else pd.Series(np.nan, index=df.index)
    )

    # ── Conflict_Geometry (g) ───────────────────────────────────────────────────
    pi = get("Conflict_Geometry")
    df["Conflict_Geometry"] = pi if pi is not None else pd.Series(np.nan, index=df.index)

    return df


def bin_continuous_abstract(df: pd.DataFrame) -> pd.DataFrame:
    """
    For continuous abstract variables (Sensor_Perception, Surface_Traction, Lateral_Stability),
    bin the [0,5] float values into integer levels 0..5 using floor(x) clamped to [0,5].
    Conflict_Geometry is already categorical — left as-is.
    """
    df = df.copy()
    for col in ["Sensor_Perception", "Surface_Traction", "Lateral_Stability"]:
        if col in df.columns:
            df[f"{col}_level"] = df[col].apply(
                lambda x: int(min(N_LEVELS - 1, max(0, np.floor(x)))) if not np.isnan(x) else np.nan
            )
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Filename parsing  (shared with criticality.py / coverage_analysis.py)
# ──────────────────────────────────────────────────────────────────────────────

def parse_filename(filename: str):
    """Returns (baseline_key, run_int) or (None, None)."""
    stem = Path(filename).stem
    run_match = re.search(r"_run(\d+)$", stem, re.IGNORECASE)
    if not run_match:
        return None, None
    run_num     = int(run_match.group(1))
    prefix_part = stem[: run_match.start()].lower()
    prefix_part = re.sub(r"_scenario\d+_\w+$", "", prefix_part)

    # Match longest key first to avoid bayscen matching bayscen_common
    for key in sorted(TARGET_BASELINES.keys(), key=len, reverse=True):
        if prefix_part == key or prefix_part.startswith(key):
            return key, run_num
    return None, None   # Not a target baseline → ignore


# ──────────────────────────────────────────────────────────────────────────────
# Data loading
# ──────────────────────────────────────────────────────────────────────────────

def load_run(filepath: Path) -> pd.DataFrame:
    """Load one per-run CSV, normalise column names, derive algo_safety."""
    df = pd.read_csv(filepath)
    df.columns = df.columns.str.strip()

    # Normalise collision / minttc column names
    rename = {}
    for c in df.columns:
        cl = c.lower()
        if cl == "collision":
            rename[c] = "collision"
        elif cl.replace("_", "") == "minttc":
            rename[c] = "minttc"
        elif cl == "algo_safety":
            rename[c] = "algo_safety"
    df = df.rename(columns=rename)

    # Derive algo_safety if not present (or re-derive with epsilon fix)
    if "collision" in df.columns and "minttc" in df.columns:
        collision = df["collision"].astype(bool)
        minttc    = pd.to_numeric(df["minttc"], errors="coerce")
        minttc_clean = minttc.replace(TTC_SENTINEL, np.nan)
        df["algo_safety"] = (
            collision | (minttc_clean <= TTC_THRESHOLD + TTC_EPSILON)
        ).astype(int)
    elif "algo_safety" not in df.columns:
        raise ValueError(f"Cannot determine failures in {filepath}: "
                         "need Collision+MinTTC or algo_safety columns.")

    return df


# ──────────────────────────────────────────────────────────────────────────────
# Failure rate computation per abstract variable level
# ──────────────────────────────────────────────────────────────────────────────

def failure_rate_by_level(
    df: pd.DataFrame,
    abstract_var: str,
) -> pd.DataFrame:
    """
    For one abstract variable, compute failure rate at each level.

    For continuous vars (Sensor_Perception, Surface_Traction, Lateral_Stability):
      groups by <var>_level (int 0-5).
    For Conflict_Geometry:
      groups by the raw category (c1, c2, c4).

    Returns DataFrame with columns:
      level | n_scenarios | n_failures | failure_rate
    """
    if abstract_var in ("Sensor_Perception", "Surface_Traction", "Lateral_Stability"):
        level_col = f"{abstract_var}_level"
        if level_col not in df.columns:
            return pd.DataFrame()
        grouped = df.groupby(level_col)["algo_safety"].agg(
            n_scenarios="count",
            n_failures="sum",
        ).reset_index().rename(columns={level_col: "level"})

    elif abstract_var == "Conflict_Geometry":
        if "Conflict_Geometry" not in df.columns or df["Conflict_Geometry"].isna().all():
            return pd.DataFrame()
        grouped = df.groupby("Conflict_Geometry")["algo_safety"].agg(
            n_scenarios="count",
            n_failures="sum",
        ).reset_index().rename(columns={"Conflict_Geometry": "level"})

    else:
        return pd.DataFrame()

    grouped["failure_rate"] = grouped["n_failures"] / grouped["n_scenarios"]
    return grouped


# ──────────────────────────────────────────────────────────────────────────────
# Main data collection
# ──────────────────────────────────────────────────────────────────────────────

def collect_results(
    results_dir: Path,
    suts: list[str],
    scenarios: list[int],
) -> pd.DataFrame:
    """
    Walk results_dir, load BayScen + BayScen-Common run files,
    compute abstract variables, compute failure rates per level,
    return a long-format DataFrame with one row per
    (sut, scenario, baseline, run, abstract_var, level).
    """
    records = []

    for sut in suts:
        for scen_num in scenarios:
            scen_folder = SCENARIO_FOLDERS.get(scen_num)
            folder = results_dir / sut / scen_folder
            if not folder.is_dir():
                log.warning("Folder not found: %s", folder)
                continue

            abstract_vars = (
                ABSTRACT_VARS_S3 if scen_num == 3 else ABSTRACT_VARS_ALL
            )

            # Index files for this folder
            file_index: dict = {}   # baseline → {run_num: Path}
            for fpath in glob.glob(str(folder / "*.csv")):
                baseline, run_num = parse_filename(fpath)
                if baseline is None:
                    continue
                file_index.setdefault(baseline, {})[run_num] = Path(fpath)

            for baseline, run_files in file_index.items():
                label = TARGET_BASELINES[baseline]
                runs_found = sorted(run_files.keys())

                if len(runs_found) < 3:
                    log.warning(
                        "[WARN] %s | Scen %d | %s – only %d run(s) %s",
                        sut, scen_num, baseline, len(runs_found), runs_found,
                    )

                for run_num, fpath in sorted(run_files.items()):
                    log.info(
                        "  %-12s | Scen %d | %-16s | run%d → %s",
                        sut, scen_num, label, run_num, fpath.name,
                    )
                    try:
                        df = load_run(fpath)
                    except Exception as exc:
                        log.error("    Load failed: %s", exc)
                        continue

                    # Compute abstract variable values + levels
                    df = compute_abstract_variables(df)
                    df = bin_continuous_abstract(df)

                    n_total    = len(df)
                    n_failures = int(df["algo_safety"].sum())

                    log.info(
                        "    n=%d | failures=%d (%.1f%%)",
                        n_total, n_failures, 100 * n_failures / n_total,
                    )

                    for av in abstract_vars:
                        fr = failure_rate_by_level(df, av)
                        if fr.empty:
                            log.warning("    No data for abstract var '%s'", av)
                            continue

                        for _, row in fr.iterrows():
                            records.append({
                                "sut":           sut,
                                "scenario":      scen_num,
                                "baseline":      baseline,
                                "baseline_label":label,
                                "run":           run_num,
                                "abstract_var":  av,
                                "level":         row["level"],
                                "n_scenarios":   int(row["n_scenarios"]),
                                "n_failures":    int(row["n_failures"]),
                                "failure_rate":  round(float(row["failure_rate"]), 6),
                                "failure_rate_pct": round(float(row["failure_rate"]) * 100, 4),
                                "n_total_run":   n_total,
                                "n_failures_run":n_failures,
                            })

    if not records:
        raise RuntimeError(
            "No data collected. Check folder structure, filenames, and that "
            "bayscen / bayscen_common files exist with '_runN' suffix."
        )

    return pd.DataFrame(records)


# ──────────────────────────────────────────────────────────────────────────────
# Aggregation: mean ± std across runs
# ──────────────────────────────────────────────────────────────────────────────

def aggregate_across_runs(raw: pd.DataFrame) -> pd.DataFrame:
    """
    For each (sut, scenario, baseline, abstract_var, level),
    compute mean and std of failure_rate across runs.
    """
    grp = raw.groupby(
        ["sut", "scenario", "baseline", "baseline_label", "abstract_var", "level"]
    )["failure_rate"].agg(
        failure_rate_mean="mean",
        failure_rate_std="std",
        failure_rate_min="min",
        failure_rate_max="max",
        n_runs="count",
    ).reset_index()

    grp["failure_rate_mean_pct"] = (grp["failure_rate_mean"] * 100).round(4)
    grp["failure_rate_std_pct"]  = (grp["failure_rate_std"]  * 100).round(4)
    return grp


def pivot_failure_rates(agg: pd.DataFrame) -> pd.DataFrame:
    """
    Pivot aggregated data so each row = (sut, scenario, baseline, abstract_var)
    and columns = level_0 … level_5 (or c1/c2/c4 for Conflict_Geometry).
    """
    pivot = agg.pivot_table(
        index=["sut", "scenario", "baseline_label", "abstract_var"],
        columns="level",
        values="failure_rate_mean_pct",
        aggfunc="first",
    ).reset_index()
    pivot.columns.name = None

    # Rename numeric level columns to "Level 0" … "Level 5"
    rename = {}
    for col in pivot.columns:
        if isinstance(col, (int, float, np.integer)):
            rename[col] = f"Level {int(col)}"
        elif isinstance(col, str) and re.match(r"^c\d+$", col):
            rename[col] = f"Path {col}"
    pivot = pivot.rename(columns=rename)
    return pivot


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def _apply_base_style(ws):
    """Apply header + alternating row style to any sheet."""
    hdr  = PatternFill("solid", fgColor="1F4E79")
    bscn = PatternFill("solid", fgColor="C8E6C9")
    bscn_common = PatternFill("solid", fgColor="FFF9C4")
    alt  = PatternFill("solid", fgColor="EEF2F7")
    bw   = Font(bold=True, color="FFFFFF")
    bb   = Font(bold=True)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDBDBD")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment, cell.border = hdr, bw, ctr, brd
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Detect baseline from any cell in row that contains label
        row_text = " ".join(str(c.value or "") for c in row).lower()
        is_bayscen        = "bayscen" in row_text and "common" not in row_text
        is_bayscen_common = "bayscen-common" in row_text or "bayscen_common" in row_text

        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            if is_bayscen:
                cell.fill, cell.font = bscn, bb
            elif is_bayscen_common:
                cell.fill, cell.font = bscn_common, bb
            elif ri % 2 == 0:
                cell.fill = alt

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 25)

    ws.freeze_panes = "A2"


def export_excel(
    raw: pd.DataFrame,
    agg: pd.DataFrame,
    pivot: pd.DataFrame,
    output_path: Path,
    args: argparse.Namespace,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: RawData ──────────────────────────────────────────────────
        raw_out = raw.copy()
        raw_out["baseline"] = raw_out["baseline"].map(
            lambda b: TARGET_BASELINES.get(b, b)
        )
        raw_out = raw_out.drop(columns=["baseline_label"], errors="ignore")
        raw_out.to_excel(writer, sheet_name="RawData", index=False)

        # ── Sheet 2: AggregatedByRun ──────────────────────────────────────────
        agg_out = agg.copy()
        agg_out = agg_out.drop(columns=["baseline"], errors="ignore")
        agg_out = agg_out.rename(columns={"baseline_label": "baseline"})
        agg_out.to_excel(writer, sheet_name="AggregatedByRun", index=False)

        # ── Sheet 3: FailureRate_Pivot ────────────────────────────────────────
        pivot_out = pivot.copy()
        pivot_out = pivot_out.rename(columns={"baseline_label": "baseline"})
        pivot_out.to_excel(writer, sheet_name="FailureRate_Pivot", index=False)

        # ── Sheet 4-7: One per abstract variable (BayScen focus) ─────────────
        for av in ABSTRACT_VARS_ALL:
            sub = agg[agg["abstract_var"] == av].copy()
            if sub.empty:
                continue
            sub = sub.drop(columns=["baseline"], errors="ignore")
            sub = sub.rename(columns={"baseline_label": "baseline"})
            sheet_name = av[:31]  # Excel 31-char limit
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

        # ── Sheet 8: Parameters ───────────────────────────────────────────────
        params = [
            ("results_dir",   str(args.results)),
            ("output",        str(args.output)),
            ("suts",          str(args.suts)),
            ("scenarios",     str(args.scenarios)),
            ("baselines",     "BayScen, BayScen-Common"),
            ("abstract_vars", str(ABSTRACT_VARS_ALL)),
            ("ttc_threshold", TTC_THRESHOLD),
            ("ttc_epsilon",   TTC_EPSILON),
            ("ttc_sentinel",  TTC_SENTINEL),
            ("n_levels",      N_LEVELS),
            ("level_meaning", "0=least degraded, 5=most degraded"),
        ]
        pd.DataFrame(params, columns=["parameter", "value"]).to_excel(
            writer, sheet_name="Parameters", index=False,
        )

    # ── Post-process: apply styling ───────────────────────────────────────────
    wb = load_workbook(output_path)
    for sh in wb.sheetnames:
        _apply_base_style(wb[sh])
    wb.save(output_path)
    log.info("Saved → %s", output_path)


# ──────────────────────────────────────────────────────────────────────────────
# Console summary
# ──────────────────────────────────────────────────────────────────────────────

def print_summary(agg: pd.DataFrame) -> None:
    print("\n" + "=" * 75)
    print("  RQ4 – Failure Rate (%) by Abstract Variable Level")
    print("  (mean across 3 runs | BayScen vs BayScen-Common)")
    print("=" * 75)

    for av in ABSTRACT_VARS_ALL:
        sub = agg[agg["abstract_var"] == av]
        if sub.empty:
            continue
        print(f"\n  ▸ {av}")

        pivot = sub.pivot_table(
            index=["sut", "scenario", "baseline_label"],
            columns="level",
            values="failure_rate_mean_pct",
            aggfunc="first",
        )
        pivot.index.names = ["SUT", "Scen", "Baseline"]
        pivot.columns = [f"L{c}" for c in pivot.columns]
        print(pivot.round(2).to_string())

    print()


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="RQ4 – Failure Characterization via abstract capability variables.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--results", type=Path, default=Path("simulation results"),
        help="Root folder containing Interfuser/ and Modular/ sub-folders.",
    )
    p.add_argument(
        "--output", type=Path, default=Path("results/rq4/failure_characterization.xlsx"),
        help="Output Excel file.",
    )
    p.add_argument(
        "--suts", nargs="+", default=SUTS_DEFAULT,
        help="SUTs to process.",
    )
    p.add_argument(
        "--scenarios", nargs="+", type=int, default=SCENARIOS_DEFAULT,
        help="Scenario numbers to process (1 2 3).",
    )
    return p.parse_args()


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    args = parse_args()

    log.info("=" * 65)
    log.info("RQ4 – Failure Characterization")
    log.info("=" * 65)
    log.info("  Results dir : %s", args.results)
    log.info("  SUTs        : %s", args.suts)
    log.info("  Scenarios   : %s", args.scenarios)
    log.info("  Baselines   : BayScen, BayScen-Common")
    log.info("  Output      : %s", args.output)
    log.info("=" * 65)

    if not args.results.is_dir():
        log.error("Results directory not found: %s", args.results)
        sys.exit(1)

    # ── Collect raw failure rates ─────────────────────────────────────────────
    log.info("\nCollecting failure rates …")
    raw = collect_results(args.results, args.suts, args.scenarios)
    log.info("Collected %d raw rows.", len(raw))

    # ── Aggregate across runs ─────────────────────────────────────────────────
    log.info("Aggregating across runs …")
    agg   = aggregate_across_runs(raw)
    pivot = pivot_failure_rates(agg)

    # ── Export ────────────────────────────────────────────────────────────────
    log.info("Exporting Excel …")
    export_excel(raw, agg, pivot, args.output, args)

    # ── Console summary ───────────────────────────────────────────────────────
    print_summary(agg)

    print(f"All RQ4 outputs saved to: {args.output}\n")


if __name__ == "__main__":
    main()
