"""
run_avfuzzer.py
===============
AV-FUZZER (Li et al., 2020) adapted for BayScen evaluation.

Implements the genetic adversarial search from:
  Li et al., "AV-Fuzzer: Finding Safety Violations in Autonomous Driving
  Systems", ISSRE 2020.

Supports all three NHTSA scenarios via --scenario:
  1  Vehicle-Vehicle Junction  (PathInteraction c1/c2/c4 + 8 env params)
  2  Vehicle-Cyclist Junction  (PathInteraction c1/c2/c4 + 9 env params + TimeOfDay)
  3  Vehicle-Vehicle Cut-In    (Direction left/right + 9 env params + TimeOfDay)

CHROMOSOME STRUCTURE
--------------------
  S1:  PathInteraction {c1,c2,c4} + ComboIndex {0..3} + 8 discrete env genes
  S2:  PathInteraction {c1,c2,c4} + ComboIndex {0..3} + 9 discrete env genes (+ TimeOfDay)
  S3:  Direction {left,right}                           + 9 discrete env genes (+ TimeOfDay)

GA HYPERPARAMETERS (paper defaults, Li et al. 2020)
---------------------------------------------------
  Population size : 12
  Crossover rate  : 0.4
  Mutation rate   : 0.3
  Budget          : 648 scenarios  (S1/S2) | 648 scenarios (S3, same budget)

ADAPTATION NOTES
----------------
  • CTBC-derived importance weights replaced by collision-rate fitness (eq. 6)
  • Discrete gene space (six-level grid) matching BayScen and all other baselines
  • Local fuzzer (paper Section IV-C) and random restart (paper IV-D) included
  • Auto-save checkpoint after every scenario; auto-resume on restart
  • CARLA crash detection stops the fuzzer immediately; resume cleans up

USAGE
-----
  python run_avfuzzer.py --scenario 1 --seed 42         # S1, run 1
  python run_avfuzzer.py --scenario 2 --seed 123        # S2, run 2
  python run_avfuzzer.py --scenario 3 --seed 7          # S3, run 3
  python run_avfuzzer.py --scenario 1 --resume          # resume checkpoint
  python run_avfuzzer.py --scenario 1 --dry_run 24      # GA mechanics, no CARLA

SEEDS  (3 runs per scenario): 42 → 123 → 7
"""

import subprocess, json, os, sys, random, math, copy, time, argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils


# =============================================================================
# GA HYPERPARAMETERS
# =============================================================================

POP_SIZE             = 12
CROSSOVER_RATE       = 0.4
MUTATION_RATE        = 0.3
LOCAL_FUZZ_THRESHOLD = 0.55
LOCAL_FUZZ_GENS      = 5
LOCAL_FUZZ_MUT_RATE  = 0.6
RESTART_PATIENCE     = 5
RESTART_CANDIDATES   = 500
TTC_CAP              = 10.0
SCENARIO_TIMEOUT     = 600
LOG_DIR              = "./avfuzzer_logs"


# =============================================================================
# PARAMETER SPACE (per scenario)
# =============================================================================

ENV_PARAMS_S1 = {
    "Cloudiness":            [0, 20, 40, 60, 80, 100],
    "Precipitation":         [0, 20, 40, 60, 80, 100],
    "PrecipitationDeposits": [0, 20, 40, 60, 80, 100],
    "WindIntensity":         [0, 20, 40, 60, 80, 100],
    "FogDensity":            [0, 20, 40, 60, 80, 100],
    "FogDistance":           [0, 20, 40, 60, 80, 100],
    "Wetness":               [0, 20, 40, 60, 80, 100],
    "RoadFriction":          [0.1, 0.2, 0.4, 0.6, 0.8, 1.0],
}
ENV_PARAMS_S2 = {
    **ENV_PARAMS_S1,
    "TimeOfDay": [-90, -60, -30, 0, 30, 60, 90],
}
ENV_PARAMS_S3 = ENV_PARAMS_S2   # same 9 env params

PATH_KEYS  = ["c1", "c2", "c4"]
N_COMBOS   = 4
PATH_MAP   = {
    "c1": [
        {"StartEgo": "left",  "GoalEgo": "right", "StartOther": "base",  "GoalOther": "left"},
        {"StartEgo": "left",  "GoalEgo": "right", "StartOther": "base",  "GoalOther": "right"},
        {"StartEgo": "base",  "GoalEgo": "left",  "StartOther": "left",  "GoalOther": "right"},
        {"StartEgo": "base",  "GoalEgo": "right", "StartOther": "left",  "GoalOther": "right"},
    ],
    "c2": [
        {"StartEgo": "right", "GoalEgo": "left",  "StartOther": "base",  "GoalOther": "left"},
        {"StartEgo": "right", "GoalEgo": "base",  "StartOther": "base",  "GoalOther": "left"},
        {"StartEgo": "base",  "GoalEgo": "left",  "StartOther": "right", "GoalOther": "left"},
        {"StartEgo": "base",  "GoalEgo": "left",  "StartOther": "right", "GoalOther": "base"},
    ],
    "c4": [
        {"StartEgo": "left",  "GoalEgo": "right", "StartOther": "right", "GoalOther": "base"},
        {"StartEgo": "left",  "GoalEgo": "base",  "StartOther": "right", "GoalOther": "base"},
        {"StartEgo": "right", "GoalEgo": "base",  "StartOther": "left",  "GoalOther": "right"},
        {"StartEgo": "right", "GoalEgo": "base",  "StartOther": "left",  "GoalOther": "base"},
    ],
}
DIRECTIONS = ["left", "right"]


def _get_env_params(scenario):
    return {1: ENV_PARAMS_S1, 2: ENV_PARAMS_S2, 3: ENV_PARAMS_S3}[scenario]


# =============================================================================
# INDIVIDUAL HELPERS
# =============================================================================

def random_individual(scenario):
    """Random chromosome for the given scenario."""
    env = _get_env_params(scenario)
    ind = {p: random.choice(vals) for p, vals in env.items()}
    if scenario in [1, 2]:
        ind["PathInteraction"] = random.choice(PATH_KEYS)
        ind["ComboIndex"]      = random.randint(0, N_COMBOS - 1)
    else:
        ind["Direction"] = random.choice(DIRECTIONS)
    return ind


def expand(ind, scenario):
    """Add derived route fields for junction scenarios."""
    row = dict(ind)
    if scenario in [1, 2]:
        combo = PATH_MAP[ind["PathInteraction"]][int(ind["ComboIndex"])]
        row.update(combo)
    return row


def mutate(ind, rate, scenario):
    """Per-gene mutation with neighbour-biased local search."""
    env   = _get_env_params(scenario)
    child = dict(ind)
    if scenario in [1, 2]:
        if random.random() < rate:
            child["PathInteraction"] = random.choice(PATH_KEYS)
        if random.random() < rate:
            child["ComboIndex"] = random.randint(0, N_COMBOS - 1)
    else:
        if random.random() < rate:
            child["Direction"] = random.choice(DIRECTIONS)
    for p, vals in env.items():
        if random.random() < rate:
            cur = vals.index(child[p]) if child[p] in vals else 0
            if random.random() < 0.7:
                step    = random.choice([-1, 1])
                new_idx = max(0, min(len(vals) - 1, cur + step))
                child[p] = vals[new_idx]
            else:
                child[p] = random.choice(vals)
    return child


def crossover(p1, p2, scenario):
    """One-point crossover on the full gene list."""
    env  = _get_env_params(scenario)
    if scenario in [1, 2]:
        keys = ["PathInteraction", "ComboIndex"] + list(env.keys())
    else:
        keys = ["Direction"] + list(env.keys())
    k = random.randint(1, len(keys) - 1)
    c1, c2 = dict(p1), dict(p2)
    for key in keys[k:]:
        c1[key], c2[key] = p2[key], p1[key]
    return c1, c2


def scenario_distance(ind1, ind2, scenario):
    """Normalised distance between two chromosomes."""
    env = _get_env_params(scenario)
    # Structural axis
    if scenario in [1, 2]:
        disc = float(ind1["PathInteraction"] != ind2["PathInteraction"]) + \
               float(ind1["ComboIndex"]      != ind2["ComboIndex"])
        disc /= 2.0
    else:
        disc = float(ind1["Direction"] != ind2["Direction"])
    # Environmental axis
    total = 0.0
    for p, vals in env.items():
        n  = len(vals) - 1
        i1 = vals.index(ind1[p]) if ind1[p] in vals else 0
        i2 = vals.index(ind2[p]) if ind2[p] in vals else 0
        total += ((i1 - i2) / n) ** 2
    return disc + math.sqrt(total) / math.sqrt(len(env))


# =============================================================================
# FITNESS
# =============================================================================

def calc_fitness(collision, min_ttc):
    if collision:       return 2.0
    if min_ttc <= 0.0:  return 1.8
    return max(0.0, 1.0 - min(min_ttc, TTC_CAP) / TTC_CAP)


# =============================================================================
# SELECTION
# =============================================================================

def roulette_select(population, fitnesses):
    total = sum(fitnesses)
    if total <= 0:
        return copy.copy(random.choice(population))
    r, cum = random.uniform(0, total), 0.0
    for ind, fit in zip(population, fitnesses):
        cum += fit
        if cum >= r:
            return copy.copy(ind)
    return copy.copy(population[-1])


# =============================================================================
# SIMULATION INTERFACE
# =============================================================================

def build_command(ind, scenario, output_paths):
    """Build the subprocess command for CARLA scenario runner."""
    row = expand(ind, scenario)
    if scenario in [1, 2]:
        cmd = [
            "python", "effects_coverage.py",
            "--scenario", "IntersectionScenarioZ_11",
            "--not_visualize",
            "--Activate_IntersectionScenario_Seed",
            "--IntersectionScenario_Seed", "26",
            "--use_sit_cov",
            "--reloadWorld",
            "--output",
            "--cloudiness",             str(row["Cloudiness"]),
            "--precipitation",          str(row["Precipitation"]),
            "--precipitation_deposits", str(row["PrecipitationDeposits"]),
            "--wind_intensity",         str(row["WindIntensity"]),
            "--fog_density",            str(row["FogDensity"]),
            "--fog_distance",           str(row["FogDistance"]),
            "--wetness",                str(row["Wetness"]),
            "--friction",               str(row["RoadFriction"]),
            "--start_ego",              str(row["StartEgo"]),
            "--goal_ego",               str(row["GoalEgo"]),
            "--start_other",            str(row["StartOther"]),
            "--goal_other",             str(row["GoalOther"]),
            "--PathInteraction",        str(row["PathInteraction"]),
        ]
        if scenario == 2:
            cmd += ["--sun_altitude_angle", str(row["TimeOfDay"])]
    else:  # S3 cut-in
        cmd = [
            "python", "cutin.py",
            "--not_visualize",
            "--reloadWorld",
            "--output",
            "--direction",              str(row["Direction"]),
            "--cloudiness",             str(row["Cloudiness"]),
            "--precipitation",          str(row["Precipitation"]),
            "--precipitation_deposits", str(row["PrecipitationDeposits"]),
            "--wind_intensity",         str(row["WindIntensity"]),
            "--sun_altitude_angle",     str(row["TimeOfDay"]),
            "--fog_density",            str(row["FogDensity"]),
            "--fog_distance",           str(row["FogDistance"]),
            "--wetness",                str(row["Wetness"]),
            "--friction",               str(row["RoadFriction"]),
            "--sync",   # required for InterFuser; comment out for Modular
        ]
    return cmd


def _read_last_ttc(path):
    try:
        with open(path) as f:
            data = json.load(f)
        if isinstance(data, list) and data:
            val = float(data[-1].get("min_ttc", 9999))
            return val if val < 9000 else float("inf")
    except Exception:
        pass
    return float("inf")


def _read_last_result(path):
    try:
        with open(path) as f:
            data = json.load(f)
        if isinstance(data, list) and data:
            last = data[-1]
            return (bool(last.get("collision_occurred", False)),
                    float(last.get("run_duration", 0.0)))
    except Exception:
        pass
    return False, 0.0


def _entry_count(path):
    try:
        with open(path) as f:
            d = json.load(f)
        return len(d) if isinstance(d, list) else 0
    except Exception:
        return 0


def run_scenario(ind, scenario_index, total, scenario, output_paths, dry_run=False):
    """Run one CARLA scenario. Returns (collision, min_ttc, duration)."""
    row     = expand(ind, scenario)
    env_str = "  ".join(f"{k}={v}" for k, v in row.items() if k not in
                        ("StartEgo","GoalEgo","StartOther","GoalOther"))
    print(f"  [{scenario_index:4d}/{total}]  {env_str}", end="  ->  ", flush=True)

    if dry_run:
        print("(dry run)")
        return False, float("inf"), 0.0

    ttc_path    = output_paths["min_ttc"]
    result_path = output_paths["results"]
    ttc_before    = _entry_count(ttc_path)
    result_before = _entry_count(result_path)

    try:
        subprocess.run(
            build_command(ind, scenario, output_paths),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=SCENARIO_TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        print("TIMEOUT")
        return False, float("inf"), SCENARIO_TIMEOUT
    except Exception as e:
        print(f"ERROR: {e}")
        return False, float("inf"), 0.0

    ttc_after    = _entry_count(ttc_path)
    result_after = _entry_count(result_path)

    if ttc_after <= ttc_before or result_after <= result_before:
        print(f"\n  CARLA CRASH at scenario {scenario_index}. "
              f"Restart CARLA and re-run (checkpoint saved).\n")
        sys.exit(1)

    collision, duration = _read_last_result(result_path)
    min_ttc             = _read_last_ttc(ttc_path)
    print(f"collision={collision}  TTC={min_ttc:.2f}s  t={duration:.0f}s")
    return collision, min_ttc, duration


# =============================================================================
# ROW BUILDER
# =============================================================================

def make_row(ind, scenario_index, gen_idx, source, collision, min_ttc, fit, scenario):
    row = expand(ind, scenario)
    env = _get_env_params(scenario)
    r   = {
        "ScenarioIndex": scenario_index,
        "Generation":    gen_idx,
        "Source":        source,
        "Collision":     collision,
        "MinTTC":        round(min_ttc, 3) if not math.isinf(min_ttc) else 9999,
        "Fitness":       round(fit, 5),
    }
    if scenario in [1, 2]:
        r.update({"PathInteraction": ind["PathInteraction"], "ComboIndex": ind["ComboIndex"],
                   "StartEgo": row["StartEgo"], "GoalEgo": row["GoalEgo"],
                   "StartOther": row["StartOther"], "GoalOther": row["GoalOther"]})
    else:
        r["Direction"] = row["Direction"]
    for p in env:
        r[p] = row[p]
    return r


# =============================================================================
# CHECKPOINT
# =============================================================================

def _fix_inf(v):  return "inf" if isinstance(v, float) and math.isinf(v) else v
def _unfix_inf(v): return float("inf") if v == "inf" else v
def _fix_ind(ind):   return {k: _fix_inf(v) for k, v in ind.items()} if ind else None
def _unfix_ind(ind): return {k: _unfix_inf(v) for k, v in ind.items()} if ind else None


def save_checkpoint(state, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    obj = {k: v for k, v in state.items() if k != "population"}
    obj["population"]    = [_fix_ind(i) for i in state["population"]]
    obj["global_best_ind"] = _fix_ind(state.get("global_best_ind"))
    obj["global_best_fit"] = _fix_inf(state.get("global_best_fit", float("-inf")))
    obj["history"]       = [_fix_ind(i) for i in state.get("history", [])[-300:]]
    tmp = path + ".tmp"
    with open(tmp, "w") as f: json.dump(obj, f)
    os.replace(tmp, path)


def load_checkpoint(path):
    if not os.path.exists(path): return None
    try:
        with open(path) as f: raw = json.load(f)
        raw["population"]      = [_unfix_ind(i) for i in raw["population"]]
        raw["global_best_ind"] = _unfix_ind(raw.get("global_best_ind"))
        raw["global_best_fit"] = _unfix_inf(raw.get("global_best_fit", "-inf"))
        raw["history"]         = [_unfix_ind(i) for i in raw.get("history", [])]
        return raw
    except Exception as e:
        print(f"  [Checkpoint] Read failed ({e}) — starting fresh")
        return None


# =============================================================================
# LOCAL FUZZER  (paper Section IV-C)
# =============================================================================

def local_fuzzer(seed, seed_fit, state, scenario, output_paths, ckpt_path, dry_run):
    remaining = state["total"] - state["scenario_counter"] + 1
    print(f"\n  -- Local Fuzzer (seed_fit={seed_fit:.4f}, remaining={remaining}) --")

    pop      = [mutate(seed, LOCAL_FUZZ_MUT_RATE, scenario) for _ in range(POP_SIZE - 1)]
    pop.insert(0, copy.copy(seed))
    best_ind, best_fit = copy.copy(seed), seed_fit

    for lg in range(LOCAL_FUZZ_GENS):
        fits = []
        for ind in pop:
            if state["scenario_counter"] > state["total"]: break
            col, ttc, dur = run_scenario(ind, state["scenario_counter"], state["total"],
                                         scenario, output_paths, dry_run)
            f = calc_fitness(col, ttc)
            fits.append(f)
            row = make_row(ind, state["scenario_counter"], state["gen_idx"],
                           f"local_fuzzer_g{lg+1}", col, ttc, f, scenario)
            state["log_rows"].append(row)
            if col: state["violations"].append(dict(row))
            state["scenario_counter"] += 1
            save_checkpoint(state, ckpt_path)
        if not fits: break
        bi = max(range(len(fits)), key=lambda i: fits[i])
        if fits[bi] > best_fit:
            best_fit, best_ind = fits[bi], copy.copy(pop[bi])
        new_pop = [pop[bi]]
        while len(new_pop) < POP_SIZE:
            new_pop.append(mutate(random.choice(pop), LOCAL_FUZZ_MUT_RATE, scenario))
        pop = new_pop

    print(f"  -- Local Fuzzer done (best_fit={best_fit:.4f}) --\n")
    return best_ind, best_fit


# =============================================================================
# RANDOM RESTART  (paper Section IV-D)
# =============================================================================

def random_restart_pop(history, scenario):
    candidates = [random_individual(scenario) for _ in range(RESTART_CANDIDATES)]
    if not history: return random.sample(candidates, POP_SIZE)
    def div_score(c):
        sample = history[-200:]
        return sum(scenario_distance(c, h, scenario) for h in sample) / len(sample)
    candidates.sort(key=div_score, reverse=True)
    print(f"  [Restart] top diversity: {div_score(candidates[0]):.4f}")
    return candidates[:POP_SIZE]


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_excel(log_rows, violations, ga_log, timestamp, scenario):
    path = os.path.join(LOG_DIR, f"avfuzzer_s{scenario}_{timestamp}.xlsx")
    wb   = openpyxl.Workbook()
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    viol_fill = PatternFill("solid", start_color="FFD7D7")
    alt_fill  = PatternFill("solid", start_color="EEF3FB")
    center    = Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, rows, highlight=None):
        if not rows: return
        cols = list(rows[0].keys())
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(1, ci, c)
            cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, center
        ws.freeze_panes = "A2"
        for ri, row in enumerate(rows, 2):
            fill = highlight(row, ri) if highlight else None
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(ri, ci, row.get(c, ""))
                cell.alignment = center
                if fill: cell.fill = fill

    ws1 = wb.active; ws1.title = "All_Scenarios"
    write_sheet(ws1, log_rows,
                lambda r, ri: viol_fill if r.get("Collision") else (alt_fill if ri%2==0 else None))
    ws2 = wb.create_sheet("Violations")
    write_sheet(ws2, violations, lambda r, ri: viol_fill)
    ws3 = wb.create_sheet("GA_Convergence")
    write_sheet(ws3, ga_log)

    wb.save(path)
    print(f"  Excel -> {path}")
    return path


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="AV-FUZZER baseline for BayScen evaluation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run_avfuzzer.py --scenario 1 --seed 42          # S1, run 1
  python run_avfuzzer.py --scenario 2 --seed 123         # S2, run 2
  python run_avfuzzer.py --scenario 3 --seed 7           # S3, run 3
  python run_avfuzzer.py --scenario 1 --resume           # resume checkpoint
  python run_avfuzzer.py --scenario 1 --dry_run 24       # no CARLA
        """,
    )
    parser.add_argument("--scenario",   type=int, default=1, choices=[1, 2, 3])
    parser.add_argument("--seed",       type=int, default=42,
                        help="Random seed (42 / 123 / 7 for the three runs)")
    parser.add_argument("--total",      type=int, default=648,
                        help="Scenario budget (default: 648)")
    parser.add_argument("--resume",     action="store_true")
    parser.add_argument("--dry_run",    type=int, default=0, metavar="N",
                        help="Run N scenarios of GA mechanics without calling CARLA")
    parser.add_argument("--min_ttc_path",  type=str, default="",
                        help="Path to min_ttc_log.json written by Scenario Runner")
    parser.add_argument("--results_path",  type=str, default="",
                        help="Path to run_results.json written by Scenario Runner")
    args = parser.parse_args()

    random.seed(args.seed)

    ckpt_path = os.path.join(LOG_DIR, f"checkpoint_s{args.scenario}.json")
    os.makedirs(LOG_DIR, exist_ok=True)

    output_paths = {
        "min_ttc": args.min_ttc_path,
        "results": args.results_path,
    }

    # ── Resume or fresh start ─────────────────────────────────────────────────
    state = None
    if args.resume:
        state = load_checkpoint(ckpt_path)
        if state:
            print(f"  [Resume] from scenario {state['scenario_counter']}")
        else:
            print("  [Resume] no checkpoint — starting fresh")

    if state is None:
        if not args.resume and os.path.exists(ckpt_path):
            if sys.stdin.isatty():
                ans = input("\n  Checkpoint found. Resume? [y/N]: ").strip().lower()
                args.resume = (ans == "y")
            else:
                args.resume = True
                print("  [Auto-resume] checkpoint detected")
            if args.resume:
                state = load_checkpoint(ckpt_path)

    dry_run = args.dry_run > 0
    total   = args.dry_run if dry_run else args.total

    if state is None:
        state = {
            "scenario_counter": 1,
            "gen_idx":          0,
            "total":            total,
            "population":       [random_individual(args.scenario) for _ in range(POP_SIZE)],
            "global_best_ind":  None,
            "global_best_fit":  float("-inf"),
            "gens_no_improve":  0,
            "history":          [],
            "log_rows":         [],
            "violations":       [],
            "ga_log":           [],
            "timestamp":        datetime.now().strftime("%Y%m%d_%H%M%S"),
        }
    state["total"] = total

    scenario_desc = {1: "Vehicle-Vehicle Junction", 2: "Vehicle-Cyclist Junction",
                     3: "Vehicle-Vehicle Cut-In"}
    print(f"\n{'='*70}")
    print(f"  AV-FUZZER  |  S{args.scenario} — {scenario_desc[args.scenario]}")
    print(f"  Budget={total}  Pop={POP_SIZE}  mut={MUTATION_RATE}  xover={CROSSOVER_RATE}")
    print(f"  Seed={args.seed}  Mode={'DRY RUN' if dry_run else 'CARLA simulation'}")
    print(f"  Checkpoint: {ckpt_path}")
    print(f"{'='*70}\n")

    wall_start = time.time()

    while state["scenario_counter"] <= total:
        remaining    = total - state["scenario_counter"] + 1
        pop_this_gen = state["population"][:min(POP_SIZE, remaining)]

        state["gen_idx"] += 1
        gen_start = state["scenario_counter"]
        gen_end   = min(gen_start + len(pop_this_gen) - 1, total)

        print(f"\n-- Gen {state['gen_idx']}  "
              f"(scenarios {gen_start}-{gen_end}/{total}) "
              f"[violations: {len(state['violations'])}] --")

        raw_fits, gen_inds = [], []

        for ind in pop_this_gen:
            if state["scenario_counter"] > total: break
            col, ttc, dur = run_scenario(ind, state["scenario_counter"], total,
                                         args.scenario, output_paths, dry_run)
            f = calc_fitness(col, ttc)
            raw_fits.append(f)
            gen_inds.append(ind)
            state["history"].append(ind)
            row = make_row(ind, state["scenario_counter"], state["gen_idx"],
                           "ga_main", col, ttc, f, args.scenario)
            state["log_rows"].append(row)
            if col:
                state["violations"].append(dict(row))
                print(f"  * SAFETY VIOLATION at scenario {state['scenario_counter']}!")
            state["scenario_counter"] += 1
            save_checkpoint(state, ckpt_path)

        if not raw_fits: break

        # Diversity-penalised adjusted fitness
        history_prev = state["history"][:-len(raw_fits)]
        adj_fits     = []
        max_d        = math.sqrt(len(_get_env_params(args.scenario)) + 1)
        for ind, f in zip(gen_inds, raw_fits):
            if history_prev:
                sample   = history_prev[-100:]
                avg_dist = sum(scenario_distance(ind, h, args.scenario) for h in sample) / len(sample)
                penalty  = 0.1 * max(0.0, 1.0 - avg_dist / max_d)
            else:
                penalty = 0.0
            adj_fits.append(f - penalty)

        best_idx     = max(range(len(raw_fits)), key=lambda i: raw_fits[i])
        gen_best_raw = raw_fits[best_idx]
        gen_best_adj = adj_fits[best_idx]

        if gen_best_adj > state["global_best_fit"]:
            state["global_best_fit"] = gen_best_adj
            state["global_best_ind"] = copy.copy(gen_inds[best_idx])
            state["gens_no_improve"] = 0
            print(f"  ^ New global best adj={state['global_best_fit']:.4f}")
        else:
            state["gens_no_improve"] += 1

        state["ga_log"].append({
            "generation":           state["gen_idx"],
            "gen_best_raw_fitness": round(gen_best_raw, 5),
            "gen_best_adj_fitness": round(gen_best_adj, 5),
            "global_best_fitness":  round(state["global_best_fit"], 5),
            "violations_so_far":    len(state["violations"]),
            "scenarios_run":        state["scenario_counter"] - 1,
        })
        print(f"  Gen {state['gen_idx']:3d}  best_raw={gen_best_raw:.4f}  "
              f"global_best={state['global_best_fit']:.4f}  "
              f"stagnant={state['gens_no_improve']}")

        # Local fuzzer
        lf_cost = POP_SIZE * LOCAL_FUZZ_GENS
        if state["scenario_counter"] + lf_cost - 1 <= total:
            for ind, f in sorted(zip(gen_inds, raw_fits), key=lambda x: x[1], reverse=True):
                if LOCAL_FUZZ_THRESHOLD <= f < 2.0:
                    lf_best, lf_fit = local_fuzzer(ind, f, state, args.scenario,
                                                   output_paths, ckpt_path, dry_run)
                    if lf_fit > f:
                        idx = gen_inds.index(ind)
                        gen_inds[idx] = lf_best
                        adj_fits[idx] = lf_fit
                    break

        if state["scenario_counter"] > total: break

        # Random restart
        if state["gens_no_improve"] >= RESTART_PATIENCE:
            print(f"\n  [Restart] stagnant {state['gens_no_improve']} gens")
            state["population"]      = random_restart_pop(state["history"], args.scenario)
            state["gens_no_improve"] = 0
            save_checkpoint(state, ckpt_path)
            continue

        # Next generation
        new_pop = []
        if state["global_best_ind"] is not None:
            new_pop.append(copy.copy(state["global_best_ind"]))
        while len(new_pop) < POP_SIZE:
            p1 = roulette_select(gen_inds, adj_fits)
            p2 = roulette_select(gen_inds, adj_fits)
            c1, c2 = crossover(p1, p2, args.scenario) if random.random() < CROSSOVER_RATE \
                     else (copy.copy(p1), copy.copy(p2))
            new_pop.append(mutate(c1, MUTATION_RATE, args.scenario))
            if len(new_pop) < POP_SIZE:
                new_pop.append(mutate(c2, MUTATION_RATE, args.scenario))
        state["population"] = new_pop
        save_checkpoint(state, ckpt_path)

    # Final export
    wall_time = time.time() - wall_start
    total_run = state["scenario_counter"] - 1

    if not dry_run:
        excel_path = export_excel(state["log_rows"], state["violations"],
                                  state["ga_log"], state["timestamp"], args.scenario)
        with open(os.path.join(LOG_DIR, f"violations_s{args.scenario}_{state['timestamp']}.json"),
                  "w") as f:
            json.dump(state["violations"], f, indent=2)

    print(f"\n{'='*70}")
    print(f"  AV-FUZZER S{args.scenario} complete")
    print(f"  Scenarios : {total_run}  |  Wall: {wall_time/60:.1f} min")
    print(f"  Violations: {len(state['violations'])}  |  "
          f"Best fitness: {state['global_best_fit']:.4f}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
